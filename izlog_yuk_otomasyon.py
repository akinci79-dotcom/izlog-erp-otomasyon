import os
import shutil
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from urllib.parse import urlsplit
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

import ayarlar
from oracle_okuyucu import (
    kaynak_yuk_verilerini_getir,
    yeni_kayitlari_veritabaninda_guncelle,
    yuk_goods_id_getir,
)

# NOT: Aktif geliştirme/hata ayıklama sırasında HER başarılı satırda ek
# "sağlama" ekran görüntüsü almak faydalıydı, ama normal çalışmada klasörü
# hızla debug_*.png dosyalarıyla dolduruyor (bkz. ayarlar.example.py'deki
# TESHIS_EKRAN_GORUNTUSU_AL notu). Varsayılan olarak KAPALI -- sadece gerçek
# bir hata/istisna oluştuğunda alınan ekran görüntüleri (bunlar zaten nadir
# ve teşhis için gereklidir) bu bayraktan ETKİLENMİYOR, her zaman alınıyor.
TESHIS_EKRAN_GORUNTUSU_AL = bool(getattr(ayarlar, "TESHIS_EKRAN_GORUNTUSU_AL", False))

# --- YARDIMCI FONKSİYONLAR ---
def devexpress_tarih_yaz(sayfa, selector, tarih_metni):
    """
    DevExpress maskeli tarih alanını temizler ve yeni tarihi yazar.

    NEDEN BÖYLE: `Home` -> `Shift+End` -> `Backspace` bazı DevExpress maskeli
    editörlerde metin seçimini normal bir textbox gibi desteklemiyor (canlı
    testte imleç sadece mevcut tarihin sonuna gidip bir alt alana geçti,
    mevcut tarih SİLİNMEDİ). Bu yüzden önce `Ctrl+A` + `Delete` deneniyor
    (çoğu editörde çalışır); silme işe yaramadıysa (alan hâlâ eski değeri
    taşıyorsa) alanın GERÇEK mevcut uzunluğuna göre hesaplanan sayıda
    `Backspace` ile garantili bir temizlik yapılıyor (sabit sayı varsaymak
    yerine).
    """
    sayfa.wait_for_selector(selector, state="visible", timeout=15000)
    sayfa.click(selector)
    sayfa.keyboard.press("Control+A")
    sayfa.keyboard.press("Delete")
    sayfa.wait_for_timeout(200)

    try:
        kalan = sayfa.input_value(selector) or ""
    except Exception:
        kalan = ""

    if kalan.strip():
        sayfa.click(selector)
        sayfa.keyboard.press("End")
        for _ in range(len(kalan) + 4):
            sayfa.keyboard.press("Backspace")

    sayfa.type(selector, tarih_metni, delay=50)
    sayfa.press(selector, "Enter")
    sayfa.wait_for_timeout(500)
    sayfa.keyboard.press("Tab")
    sayfa.wait_for_timeout(500)


def _agsakinligini_bekle(sayfa, timeout=10000, yedek_bekleme=1500):
    """
    wait_for_load_state("networkidle") DevExpress'in arka planda sürekli
    keep-alive/heartbeat isteği attığı ekranlarda hiçbir zaman "idle" olmayıp
    timeout ile patlayabilir. Bu sarmalayıcı, timeout olursa akışı durdurmak
    yerine sabit bir yedek bekleme ile devam eder (eski davranışa güvenli düşüş).
    """
    try:
        sayfa.wait_for_load_state("networkidle", timeout=timeout)
    except Exception:
        sayfa.wait_for_timeout(yedek_bekleme)


def _tutar_adaylarini_ayikla(metin):
    """Bir metin içindeki Türkçe formatlı (1.234,56) tüm tutar adaylarını Decimal olarak döndürür."""
    adaylar = re.findall(r"-?\d{1,3}(?:\.\d{3})*,\d{2}", metin)
    sonuc = []
    for aday in adaylar:
        try:
            sonuc.append(Decimal(aday.replace(".", "").replace(",", ".")))
        except InvalidOperation:
            continue
    return sonuc


def _lookup_alani_dogrula(sayfa, selector, beklenen_deger, kaynak_yuk_no, alan_adi, hata_dosya_onek):
    """
    DevExpress lookup/arama kutusuna yazılan bir değerin GERÇEKTEN kalıcı
    olduğunu doğrular -- yalnızca "boş değil" kontrolü YETERSİZ kaldı: canlı
    testte alan yazıldıktan hemen sonra dolu görünüyordu ama kısa bir süre
    sonra ERP'nin arka plan doğrulaması tarafından eski varsayılan değere
    ("Navlun"/"NAVLUN") sıfırlanıyordu. Bu yüzden burada İKİ AŞAMALI kontrol
    yapılıyor: (1) hemen sonra dolu mu, (2) fazladan bekleme sonrası hâlâ
    beklenen değeri içeriyor mu (case-insensitive substring karşılaştırması).
    İkisi de geçerse yazılan değeri döndürür; geçmezse net bir hata verir.
    """
    try:
        ilk_deger = (sayfa.input_value(selector) or "").strip()
    except Exception:
        ilk_deger = ""

    if not ilk_deger:
        try:
            sayfa.screenshot(path=f"{hata_dosya_onek}_{kaynak_yuk_no}.png")
        except Exception:
            pass
        raise RuntimeError(
            f"[{kaynak_yuk_no}] HATA: '{alan_adi}' alanı seçilemedi (beklenen: '{beklenen_deger}', "
            f"alanda görülen: '{ilk_deger}'). Bu satırın işlenmesi durduruldu. "
            f"Ekran görüntüsü: {hata_dosya_onek}_{kaynak_yuk_no}.png"
        )

    # İkinci aşama: biraz daha bekleyip HÂLÂ doğru mu diye tekrar kontrol et
    # -- ERP'nin arka plan doğrulamasının geri alması ihtimaline karşı.
    sayfa.wait_for_timeout(1500)
    try:
        son_deger = (sayfa.input_value(selector) or "").strip()
    except Exception:
        son_deger = ""

    if beklenen_deger.strip().upper() not in son_deger.upper():
        try:
            sayfa.screenshot(path=f"{hata_dosya_onek}_geri_donus_{kaynak_yuk_no}.png")
        except Exception:
            pass
        raise RuntimeError(
            f"[{kaynak_yuk_no}] HATA: '{alan_adi}' alanına '{beklenen_deger}' yazıldı ama kısa süre "
            f"sonra ERP tarafından '{son_deger}' değerine GERİ DÖNDÜRÜLDÜ (muhtemelen ERP bu metni "
            f"geçerli bir liste öğesiyle eşleştiremedi). Bu satırın işlenmesi durduruldu. "
            f"Ekran görüntüsü: {hata_dosya_onek}_geri_donus_{kaynak_yuk_no}.png"
        )

    return son_deger


def _satirda_tutar_var_mi(satir_metni, hedef_tutar, tolerans=Decimal("0.01")):
    """
    Bir grid satırının metninde, hedef tutara (kuruş toleransıyla) eşit bir
    tutar olup olmadığını kontrol eder.

    NEDEN BÖYLE: Fatura Seç (LOV) penceresinde satırı tek bir CSS string'iyle
    ("78.279,00" gibi tam grid formatına `has_text` ile birebir eşleştirerek)
    bulmaya çalışmak kırılgan; ERP'nin gösterdiği format ile bizim ürettiğimiz
    string arasında ufak bir fark (boşluk, TL eki, farklı yuvarlama) olursa
    satır hiç bulunamaz ve akış patlar. Bunun yerine, fatura no ile filtrelenen
    satırların METNİ Python'da ayrıştırılıp gerçek tutarla (Decimal) tolerans
    dahilinde karşılaştırılıyor.
    """
    for deger in _tutar_adaylarini_ayikla(satir_metni):
        if abs(deger - hedef_tutar) <= tolerans:
            return True
    return False


def _teshis_gorunur_popup_metni(sayfa):
    """
    Bir Playwright timeout hatasından hemen ÖNCE, ekranda kullanıcı
    etkileşimi bekleyen görünür bir DevExpress popup/mesaj kutusu olup
    olmadığını anlamaya çalışır -- varsa metnini döndürür (yoksa boş string).

    NEDEN BÖYLE: `EmptyRow_btnNew` timeout'u şimdiye kadar hep "generic"
    (sebepsiz) bir `Timeout 30000ms exceeded` olarak gözlemlendi. Olası bir
    açıklama: satır Kaydet'i tıklandıktan sonra ERP, GÖRÜNMEZ/gözden kaçan
    bir doğrulama uyarısı ("Sevk Oluştur" menüsündeki gibi özel bir
    `div.uyum-popup-menu` tarzı öğe) gösteriyor ve satır bu yüzden
    kaydedilmeden asılı kalıyor olabilir. Bu fonksiyon SADECE OKUR, hiçbir
    tıklama/etkileşim YAPMAZ -- akışı değiştirmez, sadece bir sonraki hata
    mesajına ekstra teşhis bilgisi ekler.
    """
    adaylar = [
        "div.dxpc-content:visible",
        "div.dx-popup-content:visible",
        "div[class*='MessageBox']:visible",
        "div[class*='messagebox']:visible",
        "div.uyum-popup-menu:visible",
        "table[id*='pm_']:visible",
    ]
    parcalar = []
    for secici in adaylar:
        try:
            metinler = sayfa.locator(secici).all_inner_texts()
        except Exception:
            continue
        for metin in metinler:
            metin = (metin or "").strip()
            if metin and metin not in parcalar:
                parcalar.append(metin)
    return " | ".join(parcalar)[:500]


def _emptyrow_bekle_teshisli(sayfa, kaynak_yuk_no, satir_index, op_kodu, satir_etiketi, timeout, asama,
                              grid_onek="TabControl_grd_LGoodsOpDetailCollection"):
    """
    Bir satır Kaydet edildikten sonra grid'in bir SONRAKİ satır için hazır
    olmasını bekler. Bu, İKİ farklı şekilde gerçekleşebilir (kullanıcının
    canlı gözlemine dayanan yeni bulgu): (1) `EmptyRow_btnNew` ("Yeni Satır
    Ekle") butonu tekrar görünür olur -- eskiden SADECE bu bekleniyordu, VEYA
    (2) ERP otomatik olarak bir SONRAKİ satırı hemen açar (bu durumda
    `EmptyRow_btnNew` GİZLİ/PASİF kalabilir, ama `DXEditor4_I` -- Tutar
    alanı -- zaten görünür olur). Eskiden sadece (1) beklendiği için, ERP
    (2) senaryosunu izlediğinde kod hiç gerçekleşmeyecek bir şeyi (buton
    görünür olsun) bekleyip gereksiz yere timeout'a düşüyordu. Artık İKİSİNDEN
    HERHANGİ BİRİ yeterli sayılıyor (CSS virgülüyle "veya" mantığı).

    `grid_onek`: Yük'ün "Yurtiçi Yük Tanımı" grid'i (`LGoodsOpDetailCollection`,
    varsayılan) ile Sevk'in fiyat grid'i (`LTransOpDetailCollection`) AYNI
    DevExpress id kalıbını (`{onek}_EmptyRow_btnNew` / `{onek}_DXEditor4_I`)
    kullanıyor -- bu yüzden fonksiyon genelleştirildi, Sevk fazında da satır
    bazlı Kaydet sonrası aynı güvenli bekleme kullanılabiliyor.

    Timeout olduğunda SADECE generic "Timeout ...ms exceeded" hatası fırlatmak
    yerine: satır indeksi + Operasyon Kodu + (varsa) görünür popup metni +
    satıra özel bir teşhis ekran görüntüsü ile zenginleştirilmiş bir
    RuntimeError fırlatılıyor.
    """
    try:
        sayfa.wait_for_selector(
            f"#{grid_onek}_EmptyRow_btnNew, "
            f"#{grid_onek}_DXEditor4_I",
            state="visible", timeout=timeout
        )
    except Exception as orijinal_hata:
        popup_metni = _teshis_gorunur_popup_metni(sayfa)
        teshis_dosyasi = f"debug_EMPTYROW_TIMEOUT_{satir_etiketi}_{kaynak_yuk_no}.png"
        try:
            sayfa.screenshot(path=teshis_dosyasi)
        except Exception:
            pass

        if popup_metni:
            ek_bilgi = f"Ekranda görünür bir popup/uyarı metni tespit edildi: '{popup_metni}'."
        else:
            ek_bilgi = "Ekranda bilinen bir popup/uyarı deseni tespit edilemedi (bilinen desenler yetersiz kalmış olabilir, ekran görüntüsüne bakılmalı)."

        raise RuntimeError(
            f"[{kaynak_yuk_no}] HATA: Satır {satir_index} ({op_kodu}) için '{asama}' sonrası "
            f"grid {timeout}ms içinde 'yeni satır ekle' durumuna dönmedi. {ek_bilgi} "
            f"Teşhis ekran görüntüsü: {teshis_dosyasi}. (Orijinal hata: {orijinal_hata})"
        ) from orijinal_hata


def _plaka_normalize(metin):
    """Plakayı karşılaştırma için sadeleştirir: boşluk/tire/nokta atılır, büyük harf."""
    return re.sub(r"[^A-Z0-9]", "", (metin or "").upper())


def _input_tutarini_coz(metin):
    """
    Bir input kutusundan okunan tutarı Decimal'e çevirir.
    Hem giriş formatını (`73000,00`) hem grid formatını (`73.000,00`) kabul eder.

    ÖNEMLİ: Önce TÜM metin parse edilir. Grid-aday regex'i (`1.234,56`)
    `73000,00` içinde yanlışlıkla `000,00` yakalayıp 0 döndürebilir.
    """
    metin = (metin or "").strip()
    if not metin:
        return None
    try:
        if "," in metin:
            return Decimal(metin.replace(".", "").replace(",", "."))
        return Decimal(metin.replace(",", ""))
    except InvalidOperation:
        pass
    adaylar = _tutar_adaylarini_ayikla(metin)
    if adaylar:
        return adaylar[0]
    return None


def _devexpress_alana_yaz(sayfa, selector, metin, delay=50, sonra_tus="Tab"):
    """
    DevExpress editörüne güvenli yazma: tıkla → tümünü seç → sil →
    karakter karakter type() → Tab (veya verilen tuş). `.fill()` KULLANILMAZ.
    """
    sayfa.wait_for_selector(selector, state="visible", timeout=15000)
    sayfa.click(selector, force=True)
    sayfa.keyboard.press("Control+A")
    sayfa.keyboard.press("Delete")
    sayfa.type(selector, metin, delay=delay)
    if sonra_tus:
        sayfa.press(selector, sonra_tus)


def _fiyat_satirini_duzenlemeye_ac(sayfa, grid_onek, kaynak_yuk_no, etiket):
    """
    Fiyat grid'inde düzenlenebilir satırı hazırlar. Yük tarafında doğrulanan
    desen: ERP bazen satırı otomatik açar (`DXEditor4_I` zaten görünür) —
    o durumda `EmptyRow_btnNew`'e basılmaz. Görünmüyorsa butona basılır.
    """
    tutar_sel = f"#{grid_onek}_DXEditor4_I"
    yeni_btn = f"#{grid_onek}_EmptyRow_btnNew"
    try:
        sayfa.wait_for_selector(tutar_sel, state="visible", timeout=2000)
        print(
            f"[{kaynak_yuk_no}] Bilgi: {etiket} fiyat satırı ZATEN açık görünüyor "
            f"-- 'Yeni Satır Ekle' butonuna basılmıyor."
        )
        return
    except Exception:
        pass
    sayfa.click(yeni_btn)
    sayfa.wait_for_selector(tutar_sel, state="visible", timeout=15000)


def _yuk_listesinde_ara_ve_sec(page, yuk_no):
    """
    Yük Listesi ekranına gider, referans no'ya göre arar ve satırı seçer.
    Hem Faz 1'in başındaki hem de `_kayitli_yuk_detay_formunu_ac`'ın kendi
    kendine yeter (self-contained) olması için ortak kullanılıyor -- direkt
    URL denemesi (`page.goto`) sayfayı listeden uzaklaştırdıysa, yedek
    (fallback) yönteme geçerken listeye GERİ DÖNÜLMESİ gerekiyor.
    """
    page.goto(ayarlar.ERP_YUK_LISTESI_URL)
    page.wait_for_selector("#myListPage_DXFREditorcol1_I", state="visible", timeout=15000)
    page.fill("#myListPage_DXFREditorcol1_I", yuk_no)
    page.press("#myListPage_DXFREditorcol1_I", "Enter")
    page.wait_for_timeout(2000)
    saglam_secici = f"tr.dxgvDataRow_Aqua:has-text('{yuk_no}')"
    page.wait_for_selector(saglam_secici, state="visible", timeout=15000)
    page.click(saglam_secici)
    return saglam_secici


def _kayitli_yuk_detay_formunu_direkt_url_ile_ac(page, islem_yuk_no, kaynak_yuk_no):
    """
    Kayıtlı Yük'ün İncele (Analyze) formunu Oracle'dan okunan GOODS_ID ile
    DOĞRUDAN URL üzerinden açar -- buton tıklama / id tahmini YOK.

    ✅ DOĞRULANMIŞ [kullanıcı canlı ERP'de bu URL'yi ekran görüntüsüyle
    paylaştı]: İncele ekranının adresi `GeneralCard.aspx?CommandName=
    LGoodsCollection.Analyze&ObjectId={GOODS_ID}&WinId=01` kalıbında --
    `ayarlar.ERP_YUK_LISTESI_URL` içindeki `CommandName=LGoodsCollection.
    Show` ile AYNI aile, sadece komut adı farklı. Bu, Yük Listesi'nde
    satırı arayıp "İncele" butonunu bulup tıklamaktan (yeni pencere mi/
    aynı sayfa mı belirsizliği, buton id/metin tahmini dahil) ÇOK daha
    güvenilir. Başarısız olursa (Oracle'dan GOODS_ID okunamazsa veya
    açılan sayfa beklenen alanları göstermezse) `None` döner -- çağıran
    taraf eski buton/çift-tıklama yöntemine (`_kayitli_yuk_detay_formunu_ac`)
    düşer.
    """
    try:
        goods_id = yuk_goods_id_getir(islem_yuk_no)
    except Exception as e:
        print(
            f"[{kaynak_yuk_no}] Uyarı: {islem_yuk_no} için GOODS_ID Oracle'dan "
            f"okunamadı ({e}) -- buton/çift tıklama yöntemine düşülüyor."
        )
        return None

    taban = urlsplit(ayarlar.ERP_LOGIN_URL)
    incele_url = (
        f"{taban.scheme}://{taban.netloc}/GeneralCard.aspx"
        f"?CommandName=LGoodsCollection.Analyze&ObjectId={goods_id}&WinId=01"
    )

    try:
        page.goto(incele_url)
        page.wait_for_selector("#TabControl_txt_ReferenceNo_I", state="visible", timeout=15000)
        okunan_ref = (page.input_value("#TabControl_txt_ReferenceNo_I") or "").strip()
    except Exception as e:
        try:
            page.screenshot(path=f"debug_sevk_direkt_url_hata_{kaynak_yuk_no}.png")
        except Exception:
            pass
        print(
            f"[{kaynak_yuk_no}] Uyarı: Doğrudan URL ({incele_url}) ile açılan sayfada "
            f"ReferenceNo alanı görünmedi ({e}) -- buton/çift tıklama yöntemine düşülüyor."
        )
        return None

    if islem_yuk_no not in okunan_ref:
        print(
            f"[{kaynak_yuk_no}] Uyarı: Doğrudan URL (GOODS_ID={goods_id}) ile açılan formun "
            f"referans no'su ('{okunan_ref}') beklenenle ('{islem_yuk_no}') eşleşmiyor -- "
            f"buton/çift tıklama yöntemine düşülüyor."
        )
        return None

    print(
        f"[{kaynak_yuk_no}] Bilgi: Kayıtlı Yük {islem_yuk_no} İncele formu "
        f"DOĞRUDAN URL ile açıldı (GOODS_ID={goods_id})."
    )
    return page


def _kayitli_yuk_detay_formunu_ac(page, islem_yuk_no, kaynak_yuk_no):
    """
    Liste satırı seçiliyken kayıtlı Yük detay formunu açar.

    NOT: Öncelikli yöntem `_kayitli_yuk_detay_formunu_direkt_url_ile_ac`'tır
    (doğrudan URL, buton tahmini yok) -- bu fonksiyon SADECE o başarısız
    olursa yedek olarak çağrılır. Kendi kendine yeterlidir: `page`'in hangi
    ekranda olduğuna bakmadan Yük Listesi'ne yeniden gider ve arar (direkt
    URL denemesi sayfayı listeden uzaklaştırmış olabilir).

    NEDEN: `YÜK OLUŞTU` / `HATA_SEVK` devamında kod Yük Listesi'nde yeni Yük
    satırını seçiyordu ama formu HİÇ açmıyordu; ardından detay alanına
    (`#TabControl_txt_ReferenceNo_I`) sağ tıklayıp "Sevk Oluştur" demeye
    çalışıyordu. O alan listede yok — ilk canlı Sevk denemesi Yük'ü oluşturup
    Sevk'te patlarsa bir sonraki çalıştırma bu yüzden baştan kırılıyordu.

    ❌ ÇÜRÜTÜLDÜ (v1) [DOĞRULANMIŞ, canlı testte teyit edildi]: Bu fonksiyonun
    ilk hali SADECE `#TabControl_txt_ReferenceNo_I`'nin "visible" olmasına
    bakıyordu ve bunu "form gerçekten açıldı" kabul ediyordu. Canlı testte
    bu YANLIŞ POZİTİF üretti: fonksiyon "form açık" diye döndü (hiçbir hata
    fırlatmadı), ama hemen sonrasındaki "Sevk Oluştur" sağ tık işlemi o alan
    üzerinde 30sn timeout'a düştü ve ekran görüntüsü sadece Yük Listesi'ni
    (hiçbir detay formu açılmamış) gösterdi. [VARSAYIM/TODO, henüz F12 ile
    kesin teyit edilmedi]: ASPx/DevExpress sayfalarında bu ID'li alanın bir
    kopyası DOM'da her zaman (ekran dışı/gizli ama Playwright'ın "visible"
    saydığı bir konumda) bulunabiliyor -- bu yüzden tek alan kontrolü
    yanıltıcı. Düzeltme: artık `#TabControl_txt_ReferenceNo_I` İLE BİRLİKTE
    `#btnSave_CD`'nin de görünür olması isteniyor (ikisi de sadece GERÇEK
    detay formunda olur).

    ✅ GÜNCEL BİLGİ [DOĞRULANMIŞ, kullanıcı canlı ERP ekran görüntüsüyle
    teyit etti]: "Sevk Oluştur" menüsü SADECE Yük penceresi **İncele
    modunda** açıkken çıkıyor (bu, "Düzelt" DEĞİL "İncele" butonuyla açılan
    pencere -- toolbar'da "Yeni, Düzelt, Sil, İncele, Kopya, Filtre, Ara"
    olarak görünüyor). Bu yüzden deneme sırası artık önce "İncele" butonu
    (görünür METİN ile bulunuyor, id tahmini YAPILMIYOR -- Türkçe "İ"
    karakteri bozulma riskine karşı regex joker kullanılıyor), sonra eski
    id tahminleri (`#btnEdit_CD` vb.) son çare olarak kalıyor. Her adım
    `expect_page()` ile YENİ pencere açılışını yarış durumu olmadan bekler
    (eskiden sabit 800ms sonra pencere sayısı kontrol ediliyordu -- pencere
    800ms'den yavaş açılırsa bu kaçırılabiliyordu).
    """
    saglam_secici = _yuk_listesinde_ara_ve_sec(page, islem_yuk_no)
    page.wait_for_timeout(500)

    def _form_gercekten_acik_mi(sayfa, timeout_ms):
        try:
            sayfa.wait_for_selector("#TabControl_txt_ReferenceNo_I", state="visible", timeout=timeout_ms)
            sayfa.wait_for_selector("#btnSave_CD", state="visible", timeout=1500)
            return True
        except Exception:
            return False

    def _eylemden_sonra_formu_bul(eylem, etiket):
        yeni = None
        try:
            with page.context.expect_page(timeout=4000) as pencere_bilgisi:
                eylem()
            yeni = pencere_bilgisi.value
        except Exception:
            yeni = None

        if yeni is not None and _form_gercekten_acik_mi(yeni, 15000):
            print(
                f"[{kaynak_yuk_no}] Bilgi: Kayıtlı Yük {islem_yuk_no} detay formu "
                f"yeni pencerede açıldı ({etiket})."
            )
            return yeni

        if _form_gercekten_acik_mi(page, 5000):
            print(
                f"[{kaynak_yuk_no}] Bilgi: Kayıtlı Yük {islem_yuk_no} detay formu "
                f"aynı sayfada açıldı ({etiket})."
            )
            return page

        return None

    # ÖNCELİK: "İncele" butonu -- kullanıcı canlı ERP'de "Sevk Oluştur"
    # menüsünün SADECE İncele modunda açılan pencerede çıktığını teyit etti.
    # Metin ile bulunuyor (id tahmini YOK); "İ" harfi joker (".") ile
    # eşleştiriliyor (dosya encoding bozulma riski, bkz. diğer Türkçe
    # karakterli seçicilerdeki aynı savunma).
    try:
        incele_adayi = page.locator("a, span, div").filter(
            has_text=re.compile(r"^\s*.ncele\s*$")
        ).first
        if incele_adayi.count() > 0:
            sonuc = _eylemden_sonra_formu_bul(
                lambda: incele_adayi.click(force=True, timeout=3000), "İncele butonu (metin ile)"
            )
            if sonuc is not None:
                return sonuc
    except Exception:
        pass

    for btn_id, etiket in (
        ("#btnEdit_CD", "Düzelt butonu #btnEdit_CD"),
        ("#btnUpdate_CD", "Düzelt butonu #btnUpdate_CD"),
        ("#btnOpen_CD", "Aç butonu #btnOpen_CD"),
        ("#btnView_CD", "İncele butonu #btnView_CD"),
        ("#btnExamine_CD", "İncele butonu #btnExamine_CD"),
    ):
        try:
            btn = page.locator(btn_id)
            if btn.count() == 0:
                continue
        except Exception:
            continue

        sonuc = _eylemden_sonra_formu_bul(
            lambda b=btn: b.first.click(force=True, timeout=3000), etiket
        )
        if sonuc is not None:
            return sonuc

    # Son çare: çift tıklama.
    sonuc = _eylemden_sonra_formu_bul(
        lambda: page.dblclick(saglam_secici), "liste satırına çift tıklama"
    )
    if sonuc is not None:
        return sonuc

    try:
        page.screenshot(path=f"debug_sevk_yuk_formu_acilamadi_{kaynak_yuk_no}.png")
    except Exception:
        pass
    raise RuntimeError(
        f"[{kaynak_yuk_no}] HATA: YÜK OLUŞTU/HATA_SEVK devamında kayıtlı Yük "
        f"{islem_yuk_no} detay formu açılamadı (Düzelt/Aç butonları ve çift "
        f"tıklama denendi, hiçbiri #TabControl_txt_ReferenceNo_I + #btnSave_CD "
        f"ikilisini birlikte görünür yapmadı). 'Sevk Oluştur' için detay formu "
        f"şart. Lütfen ERP'de bu satırı seçip 'Düzelt' butonuna elle tıklayın: "
        f"yeni pencere mi açılıyor, aynı sayfa mı değişiyor, yoksa hiç mi "
        f"tepki vermiyor? Ekran görüntüsü: debug_sevk_yuk_formu_acilamadi_{kaynak_yuk_no}.png"
    )


def _sevk_formunun_acilmasini_bekle(sayfa, kaynak_yuk_no):
    """
    'Sevk Oluştur' tıklamasından sonra Sevk formunun gerçekten yüklendiğini
    plaka alanının görünmesiyle doğrular. Timeout olursa görünür popup
    metni + ekran görüntüsü ile net hata verir (onay diyaloğu kaçmasın diye).
    """
    try:
        sayfa.wait_for_selector(
            "#TabControl_bte_TractorSerialNoPlateNo_I", state="visible", timeout=20000
        )
    except Exception as orijinal_hata:
        popup_metni = _teshis_gorunur_popup_metni(sayfa)
        teshis_dosyasi = f"debug_sevk_form_acilmadi_{kaynak_yuk_no}.png"
        try:
            sayfa.screenshot(path=teshis_dosyasi)
        except Exception:
            pass
        if popup_metni:
            ek = f"Ekranda görünür bir popup/uyarı metni tespit edildi: '{popup_metni}'."
        else:
            ek = (
                "Ekranda bilinen bir popup/uyarı deseni yok; 'Sevk Oluştur' "
                "sonrası form hiç açılmamış veya onay diyaloğu kaçırılmış olabilir."
            )
        raise RuntimeError(
            f"[{kaynak_yuk_no}] HATA: 'Sevk Oluştur' tıklandı ama Sevk formu "
            f"(plaka alanı) 20sn içinde görünmedi. {ek} "
            f"Ekran görüntüsü: {teshis_dosyasi}. (Orijinal hata: {orijinal_hata})"
        ) from orijinal_hata


# ==========================================
# UYUMSOFT ERP PLAYWRIGHT OPERASYONU V4.1 (FATURA TUTAR EŞLEŞTİRME DÜZELTMESİ)
# ==========================================
def uyumsoft_islemlerini_yap(page, kaynak_yuk_no, plaka, sevk_alis_fiyati, oracle_data, mevcut_durum, onceki_yeni_yuk_no, checkpoint_kaydet=None):

    atla_faz_123 = False
    islem_yuk_no = kaynak_yuk_no
    aktif_sayfa = page

    if mevcut_durum in ["YÜK OLUŞTU", "HATA_SEVK"] and onceki_yeni_yuk_no:
        atla_faz_123 = True
        islem_yuk_no = onceki_yeni_yuk_no

    # Güvenli dict.get() kullanımı
    satis_satirlari = oracle_data.get("SATIS_SATIRLARI", [])
    tum_faturalar = " + ".join([str(s.get("FATURA_NO", "")) for s in satis_satirlari if s.get("FATURA_NO") and str(s.get("FATURA_NO")).upper() != "NONE"])
    tum_fatura_tarihleri = " + ".join([str(s.get("FATURA_TARIHI", "")) for s in satis_satirlari if s.get("FATURA_TARIHI")])
    proje_kodu = oracle_data.get("PROJE_KODU", "")
    yuk_tarihi = oracle_data.get("YUK_TARIHI", "")
    saf_tarih = yuk_tarihi.replace(".", "") if yuk_tarihi else ""

    # --- YÜK SEÇİMİ (ORİJİNAL SAĞLAM OMURGA) ---
    _yuk_listesinde_ara_ve_sec(page, islem_yuk_no)

    # DERİN TEST MODU: ayarlar.py'de DRY_RUN=True VE DERIN_TEST_MODU=True ise,
    # sadece arama/seçimle sınırlı kalınmaz; Kopyalama + tüm veri girişi +
    # fatura LOV eşleştirmesi (asıl düzeltilen hata burada) GERÇEKTEN yapılır
    # (satır bazlı "Kaydet" dahil -- bu güvenlidir), ama ana "Kaydet"
    # (#btnSave_CD) butonuna KESİNLİKLE basılmaz — pencere kayıt yapılmadan
    # kapatılır. "Sevk Oluştur" adımı hiç test edilmez (gerçek bir Yük
    # referans numarasına ihtiyaç duyar).
    derin_test = bool(ayarlar.DRY_RUN) and bool(getattr(ayarlar, "DERIN_TEST_MODU", False))

    if ayarlar.DRY_RUN and not derin_test:
        return {"durum": "DRY_RUN BAŞARILI", "yeni_yuk_no": None, "yeni_sevk_no": None, "proje": proje_kodu, "fatura_no": tum_faturalar, "fatura_tarihi": tum_fatura_tarihleri, "tarih": yuk_tarihi, "aktif_sayfa": aktif_sayfa}

    if derin_test:
        print(f"[{kaynak_yuk_no}] ⚠️ DERİN TEST MODU AKTİF: Kopyalama ve veri girişi GERÇEKTEN yapılacak, "
              f"ama ana 'Kaydet' butonuna kesinlikle basılmayacak.")

    yeni_yuk_no = onceki_yeni_yuk_no

    # YÜK OLUŞTU / HATA_SEVK devamı: liste satırı seçili, ama Sevk menüsü
    # detay formundaki Referans No alanına sağ tık ister. Formu açmadan
    # devam etmek ilk canlı Sevk hatasından sonra yeniden denemeyi kırıyordu.
    if atla_faz_123:
        if derin_test:
            print(
                f"[{kaynak_yuk_no}] DERİN TEST: Durum '{mevcut_durum}' — Sevk fazı "
                f"bilinçli olarak atlanıyor (gerçek kayıt ister)."
            )
            return {
                "durum": "DERİN_TEST BAŞARILI",
                "yeni_yuk_no": yeni_yuk_no,
                "yeni_sevk_no": None,
                "proje": proje_kodu,
                "fatura_no": tum_faturalar,
                "fatura_tarihi": tum_fatura_tarihleri,
                "tarih": yuk_tarihi,
                "aktif_sayfa": page
            }
        print(
            f"[{kaynak_yuk_no}] Bilgi: Durum '{mevcut_durum}' — Yük fazı atlanıyor, "
            f"kayıtlı Yük {islem_yuk_no} açılarak Sevk'ten devam edilecek."
        )
        # Önce doğrudan URL yöntemi (GOODS_ID ile) denenir -- buton tıklama/
        # tahmin yok. Başarısız olursa (Oracle veya sayfa doğrulaması
        # patlarsa) eski buton/çift-tıklama yöntemine düşülür.
        aktif_sayfa = _kayitli_yuk_detay_formunu_direkt_url_ile_ac(page, islem_yuk_no, kaynak_yuk_no)
        if aktif_sayfa is None:
            aktif_sayfa = _kayitli_yuk_detay_formunu_ac(page, islem_yuk_no, kaynak_yuk_no)
        okunan_ref = (aktif_sayfa.input_value("#TabControl_txt_ReferenceNo_I") or "").strip()
        if islem_yuk_no not in okunan_ref:
            try:
                aktif_sayfa.screenshot(path=f"debug_sevk_yanlis_yuk_{kaynak_yuk_no}.png")
            except Exception:
                pass
            raise RuntimeError(
                f"[{kaynak_yuk_no}] HATA: Açılan Yük formu beklenen referans "
                f"'{islem_yuk_no}' değil (alanda görülen: '{okunan_ref}'). "
                f"Ekran görüntüsü: debug_sevk_yanlis_yuk_{kaynak_yuk_no}.png"
            )

    if not atla_faz_123:
        with page.context.expect_page() as yeni_pencere_beklentisi:
            page.click("#btnCopy_CD")

        aktif_sayfa = yeni_pencere_beklentisi.value
        aktif_sayfa.wait_for_load_state("networkidle")

        # Refactor: Yardımcı fonksiyon kullanımı
        devexpress_tarih_yaz(aktif_sayfa, "#TabControl_dte_DocDate_I", saf_tarih)

        # NOT: Sekme başlıkları regex ile ("ğ", "ü", "ı", "ç" gibi Türkçe özel
        # karakterler "." joker karakteriyle) eşleştiriliyor -- dosya ANSI/UTF-8
        # dönüşümlerinde bu karakterler bozulabiliyor (canlı testte "Diğer" ->
        # "Diger" olarak bozulup sekme bulunamadı), regex bu bozulmaya karşı
        # dayanıklı.
        aktif_sayfa.locator("span.dx-vam").filter(has_text=re.compile(r"Y.k Di.er Bilgiler")).first.click()
        aktif_sayfa.wait_for_selector("#TabControl_chk_IsGoodsInWhouse_S_D", state="visible", timeout=5000)
        aktif_sayfa.click("#TabControl_chk_IsGoodsInWhouse_S_D")

        aktif_sayfa.wait_for_timeout(500)
        aktif_sayfa.locator("span.dx-vam").filter(has_text=re.compile(r"Yurti.i Y.k Tan.m.")).first.click()
        aktif_sayfa.wait_for_selector("#TabControl_grd_LGoodsOpDetailCollection_EmptyRow_btnNew", state="visible")

        # ⚠️ KRİTİK TEŞHİS [VARSAYIM/TODO, henüz canlı ekranla kesin teyit
        # edilmedi ama güçlü şüphe var]: "Kopya" (Copy) butonu, kaynak Yük'ün
        # (zaten gerçek/kayıtlı) fiyat satırlarını da beraberinde kopyalıyor
        # olabilir. Eğer öyleyse, aşağıdaki döngü bu satırların ÜZERİNE
        # Oracle'dan okunan aynı verileri TEKRAR eklemeye çalışıyor olabilir
        # -- bu hem "aynı fatura ikinci kez bağlanmaya çalışılınca ERP'nin
        # takılması" (canlı testte gözlemlenen semptomla örtüşüyor) hem de
        # GERÇEK (DRY_RUN kapalı) çalıştırmalarda MÜKERRER KAYIT riski
        # taşıyor. Bu yüzden döngü başlamadan HEMEN ÖNCE, grid'de zaten
        # satır olup olmadığı kontrol ediliyor ve varsa NET bir uyarı +
        # ekran görüntüsü ile loglanıyor (henüz akışı DURDURMUYOR, sadece
        # teşhis amaçlı -- bir sonraki canlı testte bu uyarı çıkarsa
        # hipotez doğrulanmış olacak).
        try:
            mevcut_satir_metni = aktif_sayfa.locator(
                "#TabControl_grd_LGoodsOpDetailCollection_DXMainTable"
            ).inner_text(timeout=3000)
        except Exception:
            mevcut_satir_metni = ""

        if "Düzelt" in mevcut_satir_metni:
            print(f"[{kaynak_yuk_no}] ⚠️ UYARI: 'Yurtiçi Yük Tanımı' grid'inde döngü BAŞLAMADAN ÖNCE "
                  f"ZATEN satır(lar) var gibi görünüyor -- 'Kopya' butonunun kaynak Yük'ün mevcut fiyat "
                  f"satırlarını da kopyalamış olma ihtimali var. Bu, aşağıdaki döngünün AYNI faturayı "
                  f"ikinci kez bağlamaya çalışmasına ve/veya GERÇEK çalıştırmalarda mükerrer kayda yol "
                  f"açabilir. Lütfen kaynak Yük'ü (kopya değil, orijinali) ERP'de açıp 'Yurtiçi Yük "
                  f"Tanımı' sekmesinde zaten kayıtlı fiyat satırı olup olmadığını kontrol edin.")
            if TESHIS_EKRAN_GORUNTUSU_AL:
                try:
                    aktif_sayfa.screenshot(path=f"debug_KOPYADA_MEVCUT_SATIR_UYARISI_{kaynak_yuk_no}.png")
                except Exception:
                    pass

        # NOT: Satır bazlı "Kaydet" (a[id*='editnew']) güvenlidir -- bilinen ERP
        # kilitleme hatası, sadece ANA KAYDET (#btnSave_CD) İLE KAYDEDİLMİŞ bir
        # Yük'e daha sonra geri dönüp faturalı bir fiyat satırında tekrar işlem
        # yapmaya çalışıldığında ortaya çıkıyor (kullanıcı tarafından teyit
        # edildi). Bu otomasyonun normal akışında hiç olmuyor: her satır, o
        # Yük'ün kendi ana Kaydet'inden ÖNCE ekleniyor. Bu yüzden derin test
        # modunda tüm satırlar normal şekilde işlenir; tek fark, döngü
        # bittiğinde ana "Kaydet"e (#btnSave_CD) basılmamasıdır (aşağıda).
        # NOT: `satir_index` (1'den başlar) TEŞHİS AMAÇLI eklendi -- daha önce
        # tüm debug/hata ekran görüntüleri sadece `kaynak_yuk_no` ile
        # adlandırılıyordu (satır indeksi YOKTU). Birden fazla satış satırı
        # olduğunda (örn. aynı faturaya bağlı 2 satır), 2. satırın ekran
        # görüntüleri 1. satırınkilerin ÜZERİNE YAZILIYORDU -- bu da canlı
        # teşhiste "bu dosya hangi satıra ait?" belirsizliğine yol açtı
        # (örn. `debug_ucrettipi_operasyonkodu_sonrasi_*.png` hiç oluşmamış
        # gibi görünüp aslında 2. satırın hiç işlenemediğini mi yoksa başka
        # bir şeyin mi olduğunu ayırt etmek imkansızdı). Artık HER debug/hata
        # dosyası `satir{N}_{OPERASYON_KODU}` etiketiyle adlandırılıyor,
        # üzerine yazma tamamen ortadan kalkıyor.
        for satir_index, satis in enumerate(satis_satirlari, start=1):
            # ⚠️ KRİTİK DÜZELTME [test edilecek yeni hipotez, kullanıcının
            # canlı gözlemine dayanıyor]: Kullanıcı, bir satırı Kaydet'e
            # bastığında ERP'nin OTOMATİK OLARAK yeni (boş) bir satır açtığını
            # bildirdi -- yani "Yeni Satır Ekle" (`EmptyRow_btnNew`) butonuna
            # HER satır için manuel basmaya gerek olmayabilir, çünkü bir
            # önceki satırın Kaydet'i zaten yeni satırı otomatik açmış olabilir.
            # Eğer öyleyse, kod BİR ÖNCEKİ satırın Kaydet'inden SONRA zaten
            # açık olan bu yeni satırı fark etmeden tekrar `EmptyRow_btnNew`'e
            # basmaya çalışıyor -- bu buton, bir satır düzenleme modundayken
            # GİZLİ/PASİF olabilir, bu yüzden `wait_for_selector(...,
            # state="visible")` hiç görünmeyen bir şeyi bekleyip 90 saniye
            # sonra timeout'a düşüyor olabilir (asıl "EmptyRow_btnNew hiç
            # görünmüyor" hatasının kök nedeni tam olarak bu olabilir).
            #
            # Düzeltme: Önce DXEditor4_I (Tutar alanı) ZATEN görünür mü diye
            # bakılıyor (kısa bir timeout ile, hata fırlatmadan) -- eğer
            # görünüyorsa bir önceki satırın Kaydet'i zaten yeni satırı açmış
            # demektir, `EmptyRow_btnNew`'e HİÇ basılmıyor. Görünmüyorsa (ilk
            # satır ya da otomatik açılma olmadıysa), eskisi gibi
            # `EmptyRow_btnNew`'e basılıp yeni satırın açılması bekleniyor.
            yeni_satir_zaten_acik = False
            try:
                aktif_sayfa.wait_for_selector(
                    "#TabControl_grd_LGoodsOpDetailCollection_DXEditor4_I", state="visible", timeout=2000
                )
                yeni_satir_zaten_acik = True
                print(f"[{kaynak_yuk_no}] Bilgi: Satır {satir_index} için yeni satır ZATEN açık görünüyor "
                      f"(muhtemelen bir önceki satırın Kaydet'i otomatik açtı) -- 'Yeni Satır Ekle' "
                      f"butonuna basılmıyor.")
            except Exception:
                yeni_satir_zaten_acik = False

            if not yeni_satir_zaten_acik:
                aktif_sayfa.click("#TabControl_grd_LGoodsOpDetailCollection_EmptyRow_btnNew")
                aktif_sayfa.wait_for_selector("#TabControl_grd_LGoodsOpDetailCollection_DXEditor4_I", state="visible", timeout=15000)

            op_kodu = satis.get('OPERASYON_KODU', 'NAVLUN')
            ucret_tipi = satis.get('UCRET_TIPI', 'NAVLUN')
            satir_etiketi = f"satir{satir_index}_{op_kodu.strip().upper() or 'BILINMEYEN'}"
            print(f"[{kaynak_yuk_no}] --- Satış satırı {satir_index}/{len(satis_satirlari)} işleniyor "
                  f"(Operasyon Kodu='{op_kodu}', Ücret Tipi='{ucret_tipi}') ---")

            # NOT: Bu karşılaştırma büyük/küçük harften BAĞIMSIZ olmalı --
            # canlı testte Oracle tarafında (oracle_okuyucu.py) varsayılan
            # değer "Navlun" (baş harf büyük) olarak dönüyordu, ama burada
            # "NAVLUN" (tamamen büyük) ile karşılaştırılıyordu. Python'da bu
            # karşılaştırma harfe duyarlı olduğu için "Navlun" != "NAVLUN"
            # yanlışlıkla True çıkıp normal bir Navlun satırı "özel operasyon
            # kodu gerekiyor" sanılıyor, sonra ERP'nin arama kutusuna "Navlun"
            # yazılmaya çalışılıyor ama eşleşme bulunamayıp alan boş kalıyor
            # ve hata fırlatılıyordu.
            if op_kodu.strip().upper() != 'NAVLUN':
                # NOT: "Ücret Tipi" (DXEditor1) ve "Operasyon Kodu" (DXEditor9)
                # düz metin kutusu DEĞİL, Uyumsoft sözlük tablolarından
                # (LMSD_L_GOODSPRICE_TYPE / LMSD_L_OP_DEFINITION) beslenen
                # DevExpress ARAMA/LOOKUP kutuları -- Tutar ve Tarih alanlarında
                # daha önce görüldüğü gibi `.fill()` bu tür editörlerde
                # GÜVENİLİR DEĞİL (görsel metni yazar ama ERP'nin ihtiyaç
                # duyduğu gizli ID'yi bir öğe SEÇİLMEDEN set etmeyebilir).
                # Bu yüzden aynı sağlam desen uygulanıyor: tıkla -> temizle ->
                # karakter karakter yaz -> açılan öneri listesinin gelmesini
                # bekle -> Tab ile kutudan çık. Ardından değerin gerçekten
                # dolduğu VE (bir süre sonra) HÂLÂ doğru kaldığı doğrulanıyor.
                #
                # ÖNEMLİ DÜZELTME [DOĞRULANMIŞ, kullanıcı F12 + canlı testle
                # teyit etti]: TAM kelimeyi ("Ek_Navlun", "UĞRAMA") yazıp Tab'a
                # basmak alanı GERÇEKTEN seçilmiş yapmıyor -- ekranda kalıcı
                # olarak eski varsayılan değere ("Navlun"/"NAVLUN") geri
                # dönüyordu (`input_value()` yazdığımızı okusa da, DevExpress'in
                # kendi "seçili öğe" durumu güncellenmiyordu). Kullanıcı canlı
                # ERP'de KISA BİR ÖNEK yazıp Tab'a bastığında (örn. "ek" ->
                # "Ek_Navlun", "uğ" -> "UĞRAMA") otomatik tamamlamanın DOĞRU
                # ÇALIŞTIĞINI teyit etti -- TAM eşleşen metin yazıldığında bu
                # otomatik tamamlama/seçim mekanizması tetiklenmiyor gibi
                # görünüyor. Bu yüzden kod artık TAM değer yerine KISA BİR
                # ÖNEK yazıp Tab'a basıyor (kullanıcının elle yaptığı ile
                # birebir aynı davranış).
                onek_ucret_tipi = ucret_tipi[:3]
                aktif_sayfa.click("#TabControl_grd_LGoodsOpDetailCollection_DXEditor1_I", force=True)
                aktif_sayfa.keyboard.press("Control+A")
                aktif_sayfa.keyboard.press("Delete")
                aktif_sayfa.type("#TabControl_grd_LGoodsOpDetailCollection_DXEditor1_I", onek_ucret_tipi, delay=50)
                aktif_sayfa.wait_for_timeout(900)
                aktif_sayfa.press("#TabControl_grd_LGoodsOpDetailCollection_DXEditor1_I", "Tab")
                aktif_sayfa.wait_for_timeout(400)

                yazilan_ucret_tipi = _lookup_alani_dogrula(
                    aktif_sayfa, "#TabControl_grd_LGoodsOpDetailCollection_DXEditor1_I",
                    ucret_tipi, kaynak_yuk_no, "Ücret Tipi", f"debug_ucrettipi_hata_{satir_etiketi}"
                )

                # NOT: Aynı önek+Tab düzeltmesi Operasyon Kodu için de geçerli.
                onek_op_kodu = op_kodu[:3]
                aktif_sayfa.click("#TabControl_grd_LGoodsOpDetailCollection_DXEditor9_I", force=True)
                aktif_sayfa.keyboard.press("Control+A")
                aktif_sayfa.keyboard.press("Delete")
                aktif_sayfa.type("#TabControl_grd_LGoodsOpDetailCollection_DXEditor9_I", onek_op_kodu, delay=50)
                aktif_sayfa.wait_for_timeout(900)
                aktif_sayfa.press("#TabControl_grd_LGoodsOpDetailCollection_DXEditor9_I", "Tab")
                aktif_sayfa.wait_for_timeout(400)

                yazilan_op_kodu = _lookup_alani_dogrula(
                    aktif_sayfa, "#TabControl_grd_LGoodsOpDetailCollection_DXEditor9_I",
                    op_kodu, kaynak_yuk_no, "Operasyon Kodu", f"debug_operasyonkodu_hata_{satir_etiketi}"
                )

                print(f"[{kaynak_yuk_no}] Ücret Tipi ve Operasyon Kodu doğrulandı ve KALICI: "
                      f"Ücret Tipi='{yazilan_ucret_tipi}', Operasyon Kodu='{yazilan_op_kodu}'")

                # TEŞHİS: Ücret Tipi + Operasyon Kodu dolduktan hemen sonra
                # ekran görüntüsü al -- bir sonraki adımda (Tutar/Fatura)
                # bir şey ters giderse, bu alanların o anda GERÇEKTEN doğru
                # göründüğünü (veya görünmediğini) teyit edebilmek için.
                # Sadece TESHIS_EKRAN_GORUNTUSU_AL=True iken alınır (bkz.
                # dosya başındaki NOT) -- normal çalışmada gereksiz yere
                # klasörü doldurmaması için.
                if TESHIS_EKRAN_GORUNTUSU_AL:
                    try:
                        aktif_sayfa.screenshot(path=f"debug_ucrettipi_operasyonkodu_sonrasi_{satir_etiketi}_{kaynak_yuk_no}.png")
                    except Exception:
                        pass

            # Güvenli Decimal Çevirimi
            ham_fiyat = satis.get('SATIS_FIYATI')
            if ham_fiyat is None or str(ham_fiyat).strip() == "":
                raise ValueError(f"[{kaynak_yuk_no}] SATIS_FIYATI eksik veya geçersiz!")

            fiyat_decimal = Decimal(str(ham_fiyat))
            formatli_tutar = f"{fiyat_decimal:.2f}".replace(".", ",")

            # NOT: `.fill()` bu maskeli/formatlı DevExpress Tutar alanında
            # GÜVENİLİR DEĞİL -- canlı testte 1. satırda "şans eseri" çalıştı
            # ama 2. satırda alan boş (0,00) kaldı, akış sessizce devam edip
            # çok daha sonra anlaşılmaz bir timeout'a düştü. Tarih alanında
            # işe yarayan yöntemle aynısı uygulanıyor: tıkla -> tümünü seç ->
            # sil -> karakter karakter yaz. Ayrıca yazmanın gerçekten tuttuğu
            # doğrulanıyor; tutmadıysa hemen NET bir hata veriliyor (belirsiz
            # bir timeout yerine).
            aktif_sayfa.click("#TabControl_grd_LGoodsOpDetailCollection_DXEditor4_I", force=True)
            aktif_sayfa.keyboard.press("Control+A")
            aktif_sayfa.keyboard.press("Delete")
            aktif_sayfa.type("#TabControl_grd_LGoodsOpDetailCollection_DXEditor4_I", formatli_tutar, delay=50)
            aktif_sayfa.press("#TabControl_grd_LGoodsOpDetailCollection_DXEditor4_I", "Tab")
            aktif_sayfa.wait_for_timeout(400)

            try:
                yazilan_tutar = (aktif_sayfa.input_value("#TabControl_grd_LGoodsOpDetailCollection_DXEditor4_I") or "").strip()
            except Exception:
                yazilan_tutar = ""

            if yazilan_tutar in ("", "0,00", "0", "0,00000000", "0.00"):
                raise RuntimeError(
                    f"[{kaynak_yuk_no}] HATA: Tutar alanı doğru yazılamadı (beklenen ~'{formatli_tutar}', "
                    f"alanda görülen: '{yazilan_tutar}'). Bu satırın işlenmesi durduruldu."
                )

            # TEŞHİS: Tutar yazıldıktan hemen sonra ekran görüntüsü al -- değerin
            # sonradan (fatura bağlama veya Kaydet sırasında) sıfırlanıp
            # sıfırlanmadığını görmek için. Sadece TESHIS_EKRAN_GORUNTUSU_AL=True
            # iken alınır.
            if TESHIS_EKRAN_GORUNTUSU_AL:
                try:
                    aktif_sayfa.screenshot(path=f"debug_tutar_sonrasi_{satir_etiketi}_{kaynak_yuk_no}.png")
                except Exception:
                    pass

            # Ağ trafiğinin dinlenmesi (Sabit 3000ms yerine) - timeout olursa yedek beklemeye düşer
            _agsakinligini_bekle(aktif_sayfa, timeout=10000, yedek_bekleme=1500)

            fatura_no_raw = satis.get('FATURA_NO')
            if not fatura_no_raw or str(fatura_no_raw).strip().upper() == "NONE" or str(fatura_no_raw).strip() == "":
                aktif_sayfa.locator("a[id*='editnew']:has-text('Kaydet')").first.click(force=True)
                aktif_sayfa.wait_for_timeout(500)
                # TEŞHİS: Faturasız satırda Kaydet sonrası ekran görüntüsü.
                # Sadece TESHIS_EKRAN_GORUNTUSU_AL=True iken alınır.
                if TESHIS_EKRAN_GORUNTUSU_AL:
                    try:
                        aktif_sayfa.screenshot(path=f"debug_kaydet_sonrasi_faturasiz_{satir_etiketi}_{kaynak_yuk_no}.png")
                    except Exception:
                        pass
                _emptyrow_bekle_teshisli(
                    aktif_sayfa, kaynak_yuk_no, satir_index, op_kodu, satir_etiketi,
                    timeout=15000, asama="faturasız Kaydet"
                )
                continue

            fatura_no_str = str(fatura_no_raw).strip()

            # NOT: "3 nokta"ya tıklamadan ÖNCEKİ frame sayısını kaydediyoruz.
            # Birden fazla satış satırı işlenirken, önceki satırın LOV
            # penceresi (iframe) DOM'dan tam kaldırılmadan yeni satırın LOV'u
            # açılabiliyor -- bu durumda aşağıda `frames[-1]` almak eski,
            # kapanmakta olan frame'i yakalayabiliyor ve "Frame was detached"
            # hatasına yol açıyor (canlı testte 2. satırda gözlendi). Bu
            # yüzden aşağıda YENİ bir frame gerçekten eklenene kadar bekliyoruz.
            onceki_frame_sayisi = len(aktif_sayfa.frames)

            # ARTIK SAĞLAM: Fatura kutusunun "..." butonuna, Tab-sayarak odağı
            # tahmin etmek (ve bazen yanlış alana -- örn. Cari -- denk gelmek)
            # yerine DOĞRUDAN F12 ile bulunan sabit id'si üzerinden tıklanıyor.
            # Hangi satırda grid sütun düzeni ne olursa olsun güvenilir çalışır.
            box = None
            try:
                fatura_kutu_tablosu = aktif_sayfa.locator("#TabControl_grd_LGoodsOpDetailCollection_DXEditor29")
                fatura_kutu_tablosu.wait_for(state="visible", timeout=5000)
                box = fatura_kutu_tablosu.bounding_box()
            except Exception:
                box = None

            if not box:
                # Yedek (eski yöntem): sabit id bulunamazsa odaklanmış hücreden dene.
                print(f"[{kaynak_yuk_no}] UYARI: Fatura kutusu (DXEditor29) sabit id ile bulunamadı, "
                      f"yedek (Tab tabanlı) yönteme düşülüyor.")
                aktif_sayfa.keyboard.press("Tab")
                aktif_sayfa.wait_for_timeout(300)
                aktif_sayfa.keyboard.press("Tab")
                aktif_kutu_tablosu = aktif_sayfa.locator("*:focus").locator("xpath=ancestor::table[1]")
                box = aktif_kutu_tablosu.bounding_box()

            if box:
                btn_x = box['x'] + box['width'] - 12
                btn_y = box['y'] + (box['height'] / 2)
                aktif_sayfa.mouse.move(btn_x, btn_y)
                aktif_sayfa.wait_for_timeout(300)
                aktif_sayfa.mouse.click(btn_x, btn_y)
            else:
                raise RuntimeError(f"[{kaynak_yuk_no}] HATA: Fatura kutusu ekranda bulunamadı!")

            aktif_sayfa.wait_for_timeout(1500)

            # YENİ frame'in (LOV penceresinin) gerçekten eklendiğinden emin ol
            # -- doğrudan frames[-1] almak, eski/kapanmakta olan bir frame'i
            # yakalayıp "Frame was detached" hatasına yol açabiliyordu.
            for _ in range(20):  # ~10 saniye, 500ms aralıklarla
                if len(aktif_sayfa.frames) > onceki_frame_sayisi:
                    break
                aktif_sayfa.wait_for_timeout(500)
            else:
                raise RuntimeError(f"[{kaynak_yuk_no}] HATA: Fatura LOV penceresi (yeni frame) açılmadı!")

            lov_penceresi = aktif_sayfa.frames[-1]
            lov_penceresi.wait_for_selector("#myListPage_DXFREditorcol2_I", state="visible", timeout=20000)

            # --- Fatura no + Tutar SIRAYLA (art arda, tek seferde değil) filtreleniyor ---
            # Eskiden tutar kutusuna (#myListPage_DXFREditorcol6_I) hiç yazılmıyordu
            # çünkü "3 nokta" bazen YANLIŞ ekranı (Cari Seç) açıyordu. DXEditor29
            # düzeltmesiyle artık DOĞRU ekran (Fatura Kalemleri Listesi) açılıyor
            # ve tutar filtresi normal şekilde çalışıyor -- AMA kullanıcı canlı
            # testte önemli bir detay buldu: iki kutuyu doldurup TEK Enter'a
            # basmak işe yaramıyor (DevExpress'in filtre satırı her Enter'da bir
            # postback/yenileme tetikliyor; ikinci kutuya yazılan değer bu
            # yenilemede siliniyordu). Doğru sıra: ÖNCE fatura no yaz + Enter
            # (filtre uygulansın, bekle), SONRA tutarı (aynı format, noktasız
            # virgül ondalıklı -- örn. "426,00") yaz + Enter.
            # NOT: Bekleme, iframe'in (lov_penceresi) kendi network durumuna göre
            # yapılıyor (aktif_sayfa'nın değil) -- dış sayfa hızlıca "idle"
            # sinyali verip iframe içindeki AJAX'ın bitmesini yanlışlıkla
            # beklememiş gibi görünebiliyordu. Ayrıca her filtrenin ardından
            # DevExpress'in grid'i render etmesi için sabit bir tampon bekleme
            # de ekleniyor.
            lov_penceresi.fill("#myListPage_DXFREditorcol2_I", fatura_no_str)
            lov_penceresi.press("#myListPage_DXFREditorcol2_I", "Enter")
            _agsakinligini_bekle(lov_penceresi, timeout=10000, yedek_bekleme=2000)
            lov_penceresi.wait_for_timeout(800)

            lov_penceresi.fill("#myListPage_DXFREditorcol6_I", formatli_tutar)
            lov_penceresi.press("#myListPage_DXFREditorcol6_I", "Enter")
            _agsakinligini_bekle(lov_penceresi, timeout=10000, yedek_bekleme=2000)
            lov_penceresi.wait_for_timeout(800)

            hedef_tutar = fiyat_decimal.quantize(Decimal("0.01"))

            # --- KESİN YÖNTEM: Hücre id'sine göre doğrudan hedefleme ---
            # F12 ile incelendi: LOV grid'indeki her hücrenin id'si sabit bir
            # kalıba sahip: "myListPage|{sütunİndeksi}|{satırİndeksi}" (örn.
            # DocNo hücresi için "myListPage|2|0" -- sütun 2 = DocNo, tıpkı
            # filtre kutusu #myListPage_DXFREditorcol2_I ile aynı indeks).
            # Bu id CSS sınıfından TAMAMEN bağımsız (odaklı/seçili olsun ya da
            # olmasın değişmiyor) -- önceki CSS sınıfı tahminleri (dxgvDataRow_Aqua,
            # Row_Aqua, sınıfsız <tr> taraması) bu yüzden güvenilmezdi. ERP zaten
            # fatura no + tutar filtrelerini uyguladı; sonucun İLK satırının
            # (satır indeksi 0) DocNo hücresini doğrudan bu id ile hedefliyoruz.
            hedef_hucre = lov_penceresi.locator('td[id="myListPage|2|0"]')
            try:
                hedef_hucre.wait_for(state="visible", timeout=15000)
            except Exception:
                raise RuntimeError(
                    f"[{kaynak_yuk_no}] HATA: Fatura '{fatura_no_str}' (tutar filtresi: {formatli_tutar}) "
                    f"için LOV'da hiç satır görünmedi (id='myListPage|2|0' bulunamadı)."
                )

            # İhtiyat: ilk satırın DocNo'sunun gerçekten beklenen fatura no'yu
            # içerip içermediğini kontrol et (sadece bilgilendirme, akışı durdurmuyor).
            try:
                hucre_metni = hedef_hucre.inner_text().strip()
                if fatura_no_str not in hucre_metni:
                    print(f"[{kaynak_yuk_no}] UYARI: İlk satırın DocNo hücresi ('{hucre_metni}') "
                          f"beklenen fatura no ('{fatura_no_str}') ile birebir eşleşmiyor, "
                          f"yine de devam ediliyor (ERP filtresine güveniliyor).")
            except Exception:
                pass

            hedef_hucre.click()

            aktif_sayfa.wait_for_timeout(500)
            lov_penceresi.locator("#btnChoose_CD").click(force=True)

            # NOT: Dogru id'yi tahmin etmeye dayali dogrulama denemeleri
            # (input_value, inner_text, cesitli seciciler) hepsi yanlis
            # cikti ve satiri gereksiz yere Kaydet'e basmadan durdurdu.
            # Orijinal (1. satirda CALISAN) basit yonteme donuldu: LOV
            # kapandiktan sonra sabit bir sure bekleyip dogrudan Kaydet'e
            # basmak. 1. satirda 1500ms yetiyordu; 2. satirda yetmedigi
            # gozlendigi icin sure uzatildi.
            aktif_sayfa.wait_for_timeout(3000)

            aktif_sayfa.wait_for_selector("a[id*='editnew']:has-text('Kaydet')", state="visible", timeout=15000)
            aktif_sayfa.locator("a[id*='editnew']:has-text('Kaydet')").first.click(force=True)
            aktif_sayfa.wait_for_timeout(1500)

            # TEŞHİS: Satır Kaydet'e basıldıktan hemen sonra ekran görüntüsü al.
            # Sadece TESHIS_EKRAN_GORUNTUSU_AL=True iken alınır.
            if TESHIS_EKRAN_GORUNTUSU_AL:
                try:
                    aktif_sayfa.screenshot(path=f"debug_kaydet_sonrasi_{satir_etiketi}_{kaynak_yuk_no}.png")
                except Exception:
                    pass

            # ❌ KALDIRILDI [kendi hatam, DOĞRULANMIŞ]: Burada daha önce
            # grid metninde "Vazgeç" kelimesi kalıp kalmadığını kontrol edip
            # varsa "Kaydet tıklaması etkisiz kaldı" hatası fırlatan bir
            # kontrol vardı. Bu kontrol YANLIŞTI ve YANLIŞ ALARM üretiyordu:
            # kullanıcının bildirdiği gibi ERP, Kaydet sonrası OTOMATİK
            # olarak YENİ bir boş satır açıyor -- o yeni satır da haliyle
            # "Kaydet Vazgeç" gösteriyor! Yani grid'de "Vazgeç" görmek HER
            # ZAMAN normal (yeni satır otomatik açıldığı için), bu ORİJİNAL
            # satırın hâlâ kaydedilmediği anlamına gelmiyor. Canlı testte bu
            # kontrol, satır GERÇEKTEN başarıyla kaydedilip fatura bağlanmışken
            # bile "etkisiz kaldı" diye hata verdi (ekran görüntüsü bunu net
            # gösterdi: alt satır zaten doğru şekilde kaydedilmiş görünüyordu).
            # Kaldırıldı.

            # ⚠️ GÜNCEL BULGU [DOĞRULANMIŞ, kullanıcı canlı testte teyit etti]:
            # "Kopya"nın kaynak Yük'ün mevcut fiyat satırlarını da kopyaladığı
            # teorisi YANLIŞ çıktı -- kullanıcı, satırın (fatura dahil) GERÇEKTEN
            # otomasyon tarafından oluşturulup başarıyla kaydedildiğini teyit
            # etti. Yani veri işlemleri (satır oluşturma + fatura bağlama)
            # ÇALIŞIYOR; tek sorun, Kaydet SONRASI grid'in "yeni satır ekle"
            # durumuna dönmesinin bazen 30 saniyeden UZUN sürmesi (veya hiç
            # dönmemesi). Bu artık sadece 2. satırda değil, 1. satırda da
            # (NAVLUN, önceden hep sorunsuzdu) gözlemlendi -- yani sorun belirli
            # bir satıra/operasyon koduna özgü değil, GENEL olarak "faturalı bir
            # satırın Kaydet'i sonrası ERP'nin arka plan işleminin ne kadar
            # süreceği" ile ilgili olabilir. Bu yüzden ilk basit ve ucuz deney
            # olarak bekleme süresi 30sn'den 90sn'ye çıkarıldı -- eğer sorun
            # sadece "ERP'nin normalden yavaş olması" ise bu tek başına yeterli
            # olabilir. Yetmezse, `_emptyrow_bekle_teshisli` içindeki popup
            # tespiti ve satıra özel ekran görüntüsü bir sonraki ipucunu verecek.
            #
            # ✅ TEŞHİS KATMANI [bkz. `_emptyrow_bekle_teshisli`]: Timeout olursa
            # artık: (1) satır indeksi + operasyon kodu hata mesajına ekleniyor,
            # (2) ekranda görünür bilinen bir DevExpress popup/mesaj kutusu
            # deseni varsa metni okunup hataya ekleniyor, (3) ayrı, satıra özel
            # bir teşhis ekran görüntüsü alınıyor -- HİÇBİRİ tıklama/etkileşim
            # yapmıyor, sadece OKUYOR (akışı değiştirmiyor).
            _emptyrow_bekle_teshisli(
                aktif_sayfa, kaynak_yuk_no, satir_index, op_kodu, satir_etiketi,
                timeout=90000, asama="faturalı Kaydet"
            )

        if derin_test:
            print(f"[{kaynak_yuk_no}] DERİN TEST TAMAMLANDI: Tüm satış satırları (fatura eşleştirmesi dahil) "
                  f"başarıyla işlendi. Ana 'Kaydet' butonuna KESİNLİKLE BASILMIYOR, pencere kayıt yapılmadan "
                  f"kapatılıyor.")
            try:
                aktif_sayfa.close()
            except Exception:
                pass
            return {
                "durum": "DERİN_TEST BAŞARILI",
                "yeni_yuk_no": None,
                "yeni_sevk_no": None,
                "proje": proje_kodu,
                "fatura_no": tum_faturalar,
                "fatura_tarihi": tum_fatura_tarihleri,
                "tarih": yuk_tarihi,
                "aktif_sayfa": page
            }

        aktif_sayfa.click("#btnSave_CD", force=True)

        # JS Güvenli Parametre Enjeksiyonu
        aktif_sayfa.wait_for_function('''
            (arg) => {
                let val = document.querySelector("#TabControl_txt_ReferenceNo_I").value.trim();
                return val !== "" && val.startsWith("Y-") && val !== arg;
            }
        ''', arg=kaynak_yuk_no, timeout=20000)

        yeni_yuk_no = aktif_sayfa.input_value("#TabControl_txt_ReferenceNo_I")

        if checkpoint_kaydet:
            checkpoint_kaydet(yeni_yuk_no)

    # --- FAZ 4: SEVK OLUŞTURMA ---
    # ✅ DOĞRULANMIŞ [kullanıcı canlı ERP ekran görüntüsüyle teyit etti]:
    # "Sevk Oluştur" menüsü, Yük penceresi İncele modundayken ve sağ tık
    # ReferenceNo alanı DEĞİL, pencerenin üst tarafındaki BOŞ bir alana
    # yapıldığında çıkıyor (küçük, tek satırlık bir popup olarak görünüyor).
    # ⚠️ AÇIK SORU [VARSAYIM/TODO, kullanıcıdan kesin seçici bekleniyor]:
    # Bu "boş alan"ın tam DOM konumu/id'si henüz bilinmiyor -- koordinat
    # tahmini yapmak yerine kullanıcıya soruldu (Fatura kutusundaki eski
    # "KABA KUVVET" coord tıklama deneyiminin tekrarlanmaması için). Aşağıdaki
    # satır HENÜZ eski (bilinen YANLIŞ) hedefi kullanıyor -- kesin bilgi
    # gelince güncellenecek.
    aktif_sayfa.click("#TabControl_txt_ReferenceNo_I", button="right")
    # NOT: "ş" karakteri regex joker (".") ile eşleştiriliyor -- yukarıdaki
    # sekme başlıklarındaki encoding bozulma sorununa karşı aynı savunma.
    sevk_olustur_menu = aktif_sayfa.locator("div.uyum-popup-menu span").filter(
        has_text=re.compile(r"^Sevk Olu.tur$")
    )
    sevk_olustur_menu.first.wait_for(state="visible", timeout=15000)
    sevk_olustur_menu.first.click()

    _sevk_formunun_acilmasini_bekle(aktif_sayfa, kaynak_yuk_no)
    _agsakinligini_bekle(aktif_sayfa)

    devexpress_tarih_yaz(aktif_sayfa, "#TabControl_dte_DocDate_I", saf_tarih)

    plaka_selector = "#TabControl_bte_TractorSerialNoPlateNo_I"
    beklenen_plaka = str(plaka or "").strip()
    # Yük tarafında Enter, grid satırını erken kaydetmeye çalışabiliyordu.
    # Plaka form alanı olsa da aynı riski almamak için Tab kullanılıyor.
    _devexpress_alana_yaz(aktif_sayfa, plaka_selector, beklenen_plaka, sonra_tus="Tab")
    aktif_sayfa.wait_for_timeout(1500)

    try:
        yazilan_plaka = (aktif_sayfa.input_value(plaka_selector) or "").strip()
    except Exception:
        yazilan_plaka = ""

    if not yazilan_plaka or (
        beklenen_plaka
        and _plaka_normalize(beklenen_plaka) not in _plaka_normalize(yazilan_plaka)
        and _plaka_normalize(yazilan_plaka) not in _plaka_normalize(beklenen_plaka)
    ):
        try:
            aktif_sayfa.screenshot(path=f"debug_sevk_plaka_hata_{kaynak_yuk_no}.png")
        except Exception:
            pass
        raise RuntimeError(
            f"[{kaynak_yuk_no}] HATA: Sevk formunda Plaka alanı doldurulamadı "
            f"(beklenen: '{beklenen_plaka}', alanda görülen: '{yazilan_plaka}'). "
            f"Bu alan lookup ise tam plaka + Tab yetmeyebilir — canlı ekran "
            f"görüntüsüne bakılmalı. Ekran görüntüsü: debug_sevk_plaka_hata_{kaynak_yuk_no}.png"
        )

    if TESHIS_EKRAN_GORUNTUSU_AL:
        try:
            aktif_sayfa.screenshot(path=f"debug_sevk_plaka_sonrasi_{kaynak_yuk_no}.png")
        except Exception:
            pass

    _fiyat_satirini_duzenlemeye_ac(
        aktif_sayfa, "TabControl_grd_LTransOpDetailCollection",
        kaynak_yuk_no, "Sevk"
    )

    sevk_guvenli = Decimal(str(sevk_alis_fiyati)) if sevk_alis_fiyati else Decimal('0.00')
    formatli_sevk_fiyati = f"{sevk_guvenli:.2f}".replace(".", ",")

    sevk_tutar_selector = "#TabControl_grd_LTransOpDetailCollection_DXEditor4_I"
    _devexpress_alana_yaz(aktif_sayfa, sevk_tutar_selector, formatli_sevk_fiyati, sonra_tus="Tab")
    aktif_sayfa.wait_for_timeout(400)

    try:
        yazilan_sevk_tutari = (aktif_sayfa.input_value(sevk_tutar_selector) or "").strip()
    except Exception:
        yazilan_sevk_tutari = ""

    okunan_sevk_tutari = _input_tutarini_coz(yazilan_sevk_tutari)
    if okunan_sevk_tutari is None or abs(okunan_sevk_tutari - sevk_guvenli) > Decimal("0.01"):
        try:
            aktif_sayfa.screenshot(path=f"debug_sevk_tutar_hata_{kaynak_yuk_no}.png")
        except Exception:
            pass
        raise RuntimeError(
            f"[{kaynak_yuk_no}] HATA: Sevk Fiyatı alanı doğru yazılamadı "
            f"(beklenen ~'{formatli_sevk_fiyati}', alanda görülen: '{yazilan_sevk_tutari}'). "
            f"Ekran görüntüsü: debug_sevk_tutar_hata_{kaynak_yuk_no}.png"
        )

    if TESHIS_EKRAN_GORUNTUSU_AL:
        try:
            aktif_sayfa.screenshot(path=f"debug_sevk_tutar_sonrasi_{kaynak_yuk_no}.png")
        except Exception:
            pass

    aktif_sayfa.click("a[id*='editnew']:has-text('Kaydet')", force=True)

    # NOT: Yük tarafında satır bazlı Kaydet sonrası grid'in "hazır" duruma
    # dönmesi bazen 30-90sn sürebiliyordu (bkz. `_emptyrow_bekle_teshisli`
    # ve ilgili NOT'lar). Ana Kaydet'e (#btnSave_CD) basmadan ÖNCE Sevk fiyat
    # satırının da GERÇEKTEN kaydedildiğinden emin olmak için aynı teşhisli
    # bekleme burada da kullanılıyor (eskiden sadece sabit 500ms bekleniyordu).
    _emptyrow_bekle_teshisli(
        aktif_sayfa, kaynak_yuk_no, 1, "SEVK_FIYATI", "sevk_fiyat_satiri",
        timeout=30000, asama="Sevk fiyat satırı Kaydet",
        grid_onek="TabControl_grd_LTransOpDetailCollection"
    )

    if TESHIS_EKRAN_GORUNTUSU_AL:
        try:
            aktif_sayfa.screenshot(path=f"debug_sevk_satir_kaydet_sonrasi_{kaynak_yuk_no}.png")
        except Exception:
            pass

    aktif_sayfa.click("#btnSave_CD", force=True)

    try:
        aktif_sayfa.wait_for_function('''
            () => {
                let val = document.querySelector("#TabControl_txt_TransportNo_I").value.trim();
                return val !== "" && val.startsWith("S-");
            }
        ''', timeout=20000)
    except Exception as orijinal_hata:
        popup_metni = _teshis_gorunur_popup_metni(aktif_sayfa)
        teshis_dosyasi = f"debug_sevk_kaydet_timeout_{kaynak_yuk_no}.png"
        try:
            aktif_sayfa.screenshot(path=teshis_dosyasi)
        except Exception:
            pass
        ek = (
            f"Ekranda görünür popup/uyarı: '{popup_metni}'."
            if popup_metni
            else "Ekranda bilinen bir popup/uyarı deseni yok."
        )
        raise RuntimeError(
            f"[{kaynak_yuk_no}] HATA: Sevk ana Kaydet sonrası Transport No "
            f"(S-...) 20sn içinde oluşmadı. {ek} Ekran görüntüsü: {teshis_dosyasi}. "
            f"(Orijinal hata: {orijinal_hata})"
        ) from orijinal_hata

    yeni_sevk_no = aktif_sayfa.input_value("#TabControl_txt_TransportNo_I")

    if aktif_sayfa is not page:
        try:
            aktif_sayfa.close()
        except Exception:
            pass

    return {
        "durum": "BAŞARILI",
        "yeni_yuk_no": yeni_yuk_no,
        "yeni_sevk_no": yeni_sevk_no,
        "proje": proje_kodu,
        "fatura_no": tum_faturalar,
        "fatura_tarihi": tum_fatura_tarihleri,
        "tarih": yuk_tarihi,
        "aktif_sayfa": page
    }

def main():
    excel_dosyasi = "islem_listesi.xlsx"

    if not os.path.exists(excel_dosyasi):
        print(f"HATA: '{excel_dosyasi}' dosyası bulunamadı. Lütfen dizini kontrol edin.")
        return

    yedek_isim = f"yedek_islem_listesi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    try:
        shutil.copy(excel_dosyasi, yedek_isim)
        print(f"Orijinal veri güvene alındı. Yedek: {yedek_isim}")
    except PermissionError:
        print("HATA: Excel dosyası açık! Lütfen Excel'i kapatıp otomasyonu tekrar çalıştırın.")
        return

    wb = load_workbook(excel_dosyasi)
    ws = wb.active

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, channel="chrome")
        context = browser.new_context(viewport={'width': 1920, 'height': 1080}, locale='tr-TR')
        page = context.new_page()

        print("ERP sistemine giriş yapılıyor...")
        page.goto(ayarlar.ERP_LOGIN_URL)
        page.wait_for_selector("input#tbxUsername", state="visible")

        page.fill("input#tbxUsername", ayarlar.ERP_KULLANICI)
        page.fill("input#tbxPassword", ayarlar.ERP_SIFRE)
        page.press("input#tbxPassword", "Enter")

        page.wait_for_timeout(4000)

        basarili_yeni_yukler = []
        basarili_yeni_sevkler = []

        max_row = ws.max_row
        for i in range(2, max_row + 1):
            kaynak_yuk_no = ws.cell(row=i, column=1).value
            plaka = ws.cell(row=i, column=2).value
            sevk_alis_fiyati = ws.cell(row=i, column=3).value

            mevcut_durum = ws.cell(row=i, column=10).value
            onceki_yeni_yuk_no = ws.cell(row=i, column=8).value

            if not kaynak_yuk_no:
                continue

            if mevcut_durum in ["BAŞARILI", "DRY_RUN BAŞARILI"]:
                print(f"[{kaynak_yuk_no}] Durum: {mevcut_durum} -> İşlem Atlanıyor...")
                continue

            print(f"\n--- İŞLEM BAŞLIYOR: {kaynak_yuk_no} ---")

            hata_sayfasi = page
            try:
                oracle_data = kaynak_yuk_verilerini_getir(kaynak_yuk_no)

                def checkpoint_kaydet(yeni_yuk_no_degeri):
                    ws.cell(row=i, column=8).value = yeni_yuk_no_degeri
                    ws.cell(row=i, column=10).value = "YÜK OLUŞTU"
                    wb.save(excel_dosyasi)
                    print(f"[{kaynak_yuk_no}] CHECKPOINT ✅: Yük {yeni_yuk_no_degeri} Excel'e kaydedildi → Sevk aşamasına geçiliyor.")

                sonuc = uyumsoft_islemlerini_yap(page, kaynak_yuk_no, plaka, sevk_alis_fiyati, oracle_data, mevcut_durum, onceki_yeni_yuk_no, checkpoint_kaydet)

                hata_sayfasi = sonuc["aktif_sayfa"]

                ws.cell(row=i, column=4).value = sonuc["proje"]
                ws.cell(row=i, column=5).value = sonuc["tarih"]
                ws.cell(row=i, column=6).value = sonuc["fatura_no"]
                ws.cell(row=i, column=7).value = sonuc["fatura_tarihi"]
                if sonuc["yeni_yuk_no"]: ws.cell(row=i, column=8).value = sonuc["yeni_yuk_no"]
                if sonuc["yeni_sevk_no"]: ws.cell(row=i, column=9).value = sonuc["yeni_sevk_no"]
                ws.cell(row=i, column=10).value = sonuc["durum"]

                if sonuc["durum"] == "BAŞARILI":
                    if sonuc["yeni_yuk_no"]: basarili_yeni_yukler.append(sonuc["yeni_yuk_no"])
                    if sonuc["yeni_sevk_no"]: basarili_yeni_sevkler.append(sonuc["yeni_sevk_no"])

                wb.save(excel_dosyasi)

            except Exception as e:
                hata_mesaji = str(e)
                print(f"HATA OLUŞTU [{kaynak_yuk_no}]: {hata_mesaji}")

                zaman_damgasi = datetime.now().strftime("%H%M%S")
                hata_foto = f"hata_{kaynak_yuk_no}_{zaman_damgasi}.png"

                # KRİTİK: durum, bu turda yazılmış checkpoint'e bakılarak
                # seçilir — döngü başındaki eski `mevcut_durum` değil.
                # Eskiden Yük aynı turda oluşup Sevk patlarsa HATA_YUK yazılıp
                # yeni Yük no siliniyordu; sonraki çalıştırma aynı Yük'ü
                # tekrar kopyalıyordu.
                guncel_durum = ws.cell(row=i, column=10).value
                guncel_yeni_yuk = ws.cell(row=i, column=8).value
                yuk_zaten_olusmus = (
                    guncel_durum in ["YÜK OLUŞTU", "HATA_SEVK"]
                    or bool(guncel_yeni_yuk)
                )
                ws.cell(row=i, column=10).value = "HATA_SEVK" if yuk_zaten_olusmus else "HATA_YUK"
                if not yuk_zaten_olusmus:
                    ws.cell(row=i, column=8).value = None
                ws.cell(row=i, column=11).value = hata_mesaji

                try:
                    if len(context.pages) > 1:
                        hata_sayfasi = context.pages[-1]

                    hata_sayfasi.screenshot(path=hata_foto)
                    hata_sayfasi.keyboard.press("Escape")

                    if len(context.pages) > 1:
                        hata_sayfasi.close()
                except Exception:
                    print(f"[{kaynak_yuk_no}] Uyarı: hata ekran görüntüsü/pencere kapatma başarısız; durum yine de kaydedildi.")

                wb.save(excel_dosyasi)
                print(f"Hata detayı kaydedildi. Ekran Görüntüsü: {hata_foto}")

                try:
                    page.goto(ayarlar.ERP_YUK_LISTESI_URL)
                except Exception:
                    print("Sistem başlangıç ekranına dönemedi! Güvenlik için otomasyon durduruluyor.")
                    break

        print("\nTüm Excel satırları tamamlandı.")

        if basarili_yeni_yukler or basarili_yeni_sevkler:
            print(f"\n>>> Veritabanı İz Temizliği (Toplu SQL İşlemi) Başlatılıyor... <<<")
            yeni_kayitlari_veritabaninda_guncelle(basarili_yeni_yukler, basarili_yeni_sevkler)

        print("\nERP'den güvenli çıkış yapılıyor...")
        try:
            # NOT: Önceden burada `timeout=` verilmiyordu -- eğer buton anlık
            # olarak tıklanabilir durumda değilse (örn. önceki işlemden kalan
            # bir yükleme/overlay yüzünden), Playwright'ın varsayılan
            # actionability timeout'u (30sn) devreye giriyordu ve bu hata
            # sessizce (`except: pass`) yutulduğu için tek görülen şey
            # ekranın "bir süre takılı kalıp sonra kapanması" oluyordu. Kısa
            # bir timeout ile bu gereksiz bekleme ortadan kaldırıldı.
            page.click("div.logout-main", timeout=5000)
            page.wait_for_timeout(2000)
        except Exception:
            pass

        browser.close()

if __name__ == "__main__":
    main()
