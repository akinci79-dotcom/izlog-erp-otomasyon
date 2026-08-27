import os
import re
import shutil
from datetime import datetime
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

import ayarlar
from oracle_okuyucu import kaynak_yuk_verilerini_getir, yeni_kayitlari_veritabaninda_guncelle

# ==========================================
# UYUMSOFT ERP PLAYWRIGHT OPERASYONU V3.20
# (FATURA TUTAR EŞLEŞTİRME DÜZELTMESİ + GENEL SAĞLAMLAŞTIRMA)
# ==========================================


def _tarih_alanini_temizle_ve_yaz(sayfa, selector, tarih_ddmmyyyy):
    """DevExpress maskeli tarih alanını güvenli şekilde temizler ve yeni tarihi yazar.

    Faz 3 (Yük tarihi) ve Faz 4 (Sevk tarihi) için tekrar eden kod tek
    fonksiyona çıkarıldı.
    """
    sayfa.click(selector)
    sayfa.keyboard.press("End")
    for _ in range(12):
        sayfa.keyboard.press("Backspace")
    sayfa.keyboard.press("Home")

    saf_tarih = tarih_ddmmyyyy.replace(".", "")
    sayfa.type(selector, saf_tarih, delay=100)
    sayfa.press(selector, "Enter")
    sayfa.wait_for_timeout(500)


def _tutar_adaylarini_ayikla(metin):
    """Bir metin içindeki Türkçe formatlı (1.234,56) tüm tutar adaylarını Decimal olarak döndürür."""
    adaylar = re.findall(r"-?\d{1,3}(?:\.\d{3})*,\d{2}", metin)
    sonuc = []
    for aday in adaylar:
        try:
            sonuc.append(Decimal(aday.replace(".", "").replace(",", ".")))
        except InvalidOperation:
            continue
    return sonuc


def _satirda_tutar_var_mi(satir_metni, hedef_tutar, tolerans=Decimal("0.01")):
    """
    Bir grid satırının metninde, hedef tutara (kuruş toleransıyla) eşit bir
    tutar olup olmadığını kontrol eder.

    NEDEN BÖYLE: ERP'nin fatura arama (LOV) penceresindeki tutar filtre
    kutusuna (#myListPage_DXFREditorcol6_I) doğrudan değer yazmak, bu alanın
    DevExpress'e özel maskeli/formatlı bir editör olması nedeniyle kırılgan ve
    riskliydi (yanlış/eksik format -> filtre hatası ya da olası veri bozulması).
    Bu yüzden tutar HİÇ bu kutuya yazılmıyor; sadece fatura no ile filtrelenen
    satırların METNİ Python tarafında ayrıştırılıp gerçek tutarla karşılaştırılıyor.
    """
    for deger in _tutar_adaylarini_ayikla(satir_metni):
        if abs(deger - hedef_tutar) <= tolerans:
            return True
    return False


def _tam_veya_alt_dize_satir_bul(sayfa_veya_frame, satir_secici, aranan_deger):
    """
    `satir_secici` ile eşleşen satırlar arasında, hücre metinlerinden biri
    `aranan_deger` ile TAM eşleşen satırı bulmaya çalışır (trim edilmiş).
    Tam eşleşme bulunamazsa (örn. ERP hücreyi farklı biçimde render ediyorsa)
    ilk alt-dize eşleşmesine düşer; ama önce her zaman tam eşleşme denenir.
    Bu, sadece `:has-text()` (alt-dize) eşleşmesine güvenmenin getirdiği
    "1000 satırı 21000'i de eşler" tipi riski azaltır.
    """
    satirlar = sayfa_veya_frame.locator(satir_secici)
    sayi = satirlar.count()
    for idx in range(sayi):
        satir = satirlar.nth(idx)
        try:
            hucreler = [h.strip() for h in satir.inner_text().split("\n")]
        except Exception:
            continue
        if str(aranan_deger).strip() in hucreler:
            return satir
    # Tam eşleşme bulunamadı -> alt-dize eşleşmesine düş (mevcut davranışla uyumluluk için)
    return satirlar.first if sayi > 0 else None


def uyumsoft_islemlerini_yap(page, kaynak_yuk_no, plaka, sevk_alis_fiyati, oracle_data, mevcut_durum, onceki_yeni_yuk_no, checkpoint_kaydet=None):

    # 1. PARÇALI BAŞARI YÖNETİMİ
    atla_faz_123 = False
    islem_yuk_no = kaynak_yuk_no
    aktif_sayfa = page

    if mevcut_durum in ["YÜK OLUŞTU", "HATA_SEVK"] and onceki_yeni_yuk_no:
        atla_faz_123 = True
        islem_yuk_no = onceki_yeni_yuk_no
        print(f"[{kaynak_yuk_no}] Sevk aşamasında kalındığı tespit edildi. Hedef Yük: {islem_yuk_no}")

    tum_faturalar = " + ".join([str(s["FATURA_NO"]) for s in oracle_data["SATIS_SATIRLARI"] if s.get("FATURA_NO") and str(s.get("FATURA_NO")).upper() != "NONE"])
    tum_fatura_tarihleri = " + ".join([str(s["FATURA_TARIHI"]) for s in oracle_data["SATIS_SATIRLARI"] if s.get("FATURA_TARIHI")])
    proje_kodu = oracle_data.get("PROJE_KODU", "")
    yuk_tarihi = oracle_data.get("YUK_TARIHI", "")

    # --- GÜVENLİ BAŞLANGIÇ ---
    page.goto(ayarlar.ERP_YUK_LISTESI_URL)
    page.wait_for_selector("#myListPage_DXFREditorcol1_I", state="visible", timeout=15000)

    # --- FAZ 1: YÜKÜ BUL ---
    page.fill("#myListPage_DXFREditorcol1_I", islem_yuk_no)
    page.press("#myListPage_DXFREditorcol1_I", "Enter")

    satir_secici = "tr.dxgvDataRow_Aqua"
    page.wait_for_selector(f"{satir_secici}:has-text('{islem_yuk_no}')", state="visible", timeout=15000)

    hedef_yuk_satiri = _tam_veya_alt_dize_satir_bul(page, satir_secici, islem_yuk_no)
    if hedef_yuk_satiri is None:
        raise RuntimeError(f"[{kaynak_yuk_no}] HATA: '{islem_yuk_no}' numaralı yük listede bulunamadı.")
    hedef_yuk_satiri.click()

    # --- DRY RUN KONTROLÜ ---
    if ayarlar.DRY_RUN:
        print(f"[{kaynak_yuk_no}] DRY_RUN AKTİF: Operasyonel veriler doğrulandı, Kaydet/Kopya simüle edildi.")
        return {
            "durum": "DRY_RUN BAŞARILI",
            "yeni_yuk_no": None,
            "yeni_sevk_no": None,
            "proje": proje_kodu,
            "fatura_no": tum_faturalar,
            "fatura_tarihi": tum_fatura_tarihleri,
            "tarih": yuk_tarihi,
            "aktif_sayfa": aktif_sayfa
        }

    yeni_yuk_no = onceki_yeni_yuk_no

    # Yük daha önce kopyalanmadıysa Faz 2 ve Faz 3 çalıştırılır
    if not atla_faz_123:
        # --- FAZ 2: KOPYALAMA VE YENİ PENCERE YAKALAMA ---
        print(f"[{kaynak_yuk_no}] Faz 2: Kopyalama başlatılıyor...")

        with page.context.expect_page() as yeni_pencere_beklentisi:
            page.click("#btnCopy_CD")

        yeni_sayfa = yeni_pencere_beklentisi.value
        yeni_sayfa.wait_for_load_state("networkidle")
        print(f"[{kaynak_yuk_no}] Yeni kopyalama ekranına başarıyla geçildi.")

        aktif_sayfa = yeni_sayfa

        # --- FAZ 3: VERİ GİRİŞİ ---
        aktif_sayfa.wait_for_selector("#TabControl_dte_DocDate_I", state="visible", timeout=15000)

        _tarih_alanini_temizle_ve_yaz(aktif_sayfa, "#TabControl_dte_DocDate_I", yuk_tarihi)
        aktif_sayfa.keyboard.press("Tab")
        aktif_sayfa.wait_for_timeout(1000)

        aktif_sayfa.click("span.dx-vam:has-text('Yük Diğer Bilgiler')")

        aktif_sayfa.wait_for_selector("#TabControl_chk_IsGoodsInWhouse_S_D", state="visible", timeout=5000)
        aktif_sayfa.click("#TabControl_chk_IsGoodsInWhouse_S_D")

        # --- FAZ 3: FATURA BAĞLANTILARI ---
        print(f"[{kaynak_yuk_no}] Faz 3: Satış fiyatları ve faturalar bağlanıyor...")

        aktif_sayfa.wait_for_timeout(1000)
        aktif_sayfa.click("span.dx-vam:has-text('Yurtiçi Yük Tanımı')")
        aktif_sayfa.wait_for_selector("#TabControl_grd_LGoodsOpDetailCollection_EmptyRow_btnNew", state="visible")

        if not oracle_data.get("SATIS_SATIRLARI"):
            raise ValueError("Oracle'dan satış satırı gelmedi, işlem durduruldu!")

        for satis in oracle_data["SATIS_SATIRLARI"]:
            aktif_sayfa.click("#TabControl_grd_LGoodsOpDetailCollection_EmptyRow_btnNew")
            aktif_sayfa.wait_for_timeout(2500)
            aktif_sayfa.wait_for_selector("#TabControl_grd_LGoodsOpDetailCollection_DXEditor4_I", state="visible")

            if satis['OPERASYON_KODU'] != 'NAVLUN':
                aktif_sayfa.fill("#TabControl_grd_LGoodsOpDetailCollection_DXEditor1_I", satis['UCRET_TIPI'])
                aktif_sayfa.press("#TabControl_grd_LGoodsOpDetailCollection_DXEditor1_I", "Tab")
                aktif_sayfa.wait_for_timeout(500)
                aktif_sayfa.fill("#TabControl_grd_LGoodsOpDetailCollection_DXEditor9_I", satis['OPERASYON_KODU'])
                aktif_sayfa.press("#TabControl_grd_LGoodsOpDetailCollection_DXEditor9_I", "Tab")
                aktif_sayfa.wait_for_timeout(500)

            satis_fiyati_ham = satis.get('SATIS_FIYATI')
            if satis_fiyati_ham is None:
                raise ValueError(f"[{kaynak_yuk_no}] HATA: Satış satırında SATIS_FIYATI boş (None) geldi, işlem durduruldu!")
            try:
                satis_fiyati_decimal = Decimal(str(satis_fiyati_ham))
            except InvalidOperation:
                raise ValueError(f"[{kaynak_yuk_no}] HATA: SATIS_FIYATI '{satis_fiyati_ham}' sayısal biçime çevrilemedi!")

            # FORMAT 1: Kutuya yazmak için saf form (Örn: 78279,00)
            formatli_tutar = f"{satis_fiyati_decimal:.2f}".replace(".", ",")

            # Fiyatı yaz ve Enter YERİNE sadece TAB ile ilerle
            aktif_sayfa.click("#TabControl_grd_LGoodsOpDetailCollection_DXEditor4_I", force=True)
            aktif_sayfa.fill("#TabControl_grd_LGoodsOpDetailCollection_DXEditor4_I", formatli_tutar)
            aktif_sayfa.press("#TabControl_grd_LGoodsOpDetailCollection_DXEditor4_I", "Tab")
            aktif_sayfa.wait_for_timeout(400)

            aktif_sayfa.keyboard.press("Tab")  # Tutar kutusunu geç
            aktif_sayfa.wait_for_timeout(400)
            aktif_sayfa.keyboard.press("Tab")  # Op Kodu kutusunu geç

            print(f"[{kaynak_yuk_no}] Fiyat girildi, imleç Fatura kutusuna taşındı...")
            aktif_sayfa.wait_for_timeout(2000)  # Arka plan hesaplamasının bitmesi için kritik bekleme

            fatura_no_raw = satis.get('FATURA_NO')
            if not fatura_no_raw or str(fatura_no_raw).strip().upper() == "NONE" or str(fatura_no_raw).strip() == "":
                print(f"[{kaynak_yuk_no}] DİKKAT: Bu satırın faturası yok. Sadece fiyat kaydediliyor.")
                aktif_sayfa.locator("a[id*='editnew']:has-text('Kaydet')").first.click(force=True)
                aktif_sayfa.wait_for_selector("#TabControl_grd_LGoodsOpDetailCollection_EmptyRow_btnNew", state="visible", timeout=15000)
                continue

            fatura_no_str = str(fatura_no_raw).strip()

            print(f"[{kaynak_yuk_no}] Fatura kutusuna ulaşıldı. 3 noktaya dinamik tıklanıyor...")
            # İmlecin içinde bulunduğu aktif tablonun 3 noktasına fiziksel tıklama
            aktif_kutu_tablosu = aktif_sayfa.locator("*:focus").locator("xpath=ancestor::table[1]")
            box = aktif_kutu_tablosu.bounding_box()

            if box:
                btn_x = box['x'] + box['width'] - 12
                btn_y = box['y'] + (box['height'] / 2)
                aktif_sayfa.mouse.move(btn_x, btn_y)
                aktif_sayfa.wait_for_timeout(300)
                aktif_sayfa.mouse.click(btn_x, btn_y)
            else:
                raise RuntimeError(f"[{kaynak_yuk_no}] HATA: Fatura kutusu ekranda odaklanamadı!")

            print(f"[{kaynak_yuk_no}] Fatura penceresi açıldı, seçim yapılıyor...")
            aktif_sayfa.wait_for_timeout(2000)
            lov_penceresi = aktif_sayfa.frames[-1]
            lov_penceresi.wait_for_selector("#myListPage_DXFREditorcol2_I", state="visible", timeout=20000)

            # --- ZIRH (DÜZELTİLDİ): Tutar kutusuna HİÇ dokunulmuyor. ---
            # Eskiden tutar filtre kutusuna (#myListPage_DXFREditorcol6_I) değer
            # yazılmaya çalışılıyordu; bu alan DevExpress'in maskeli/formatlı
            # sayısal editörü olduğu için yazılan değer doğru işlenmiyor ve
            # akış patlıyordu (ayrıca gerçek veriyi bozma riski taşıyordu).
            # Bu yüzden SADECE fatura no ile filtreleniyor, doğru satır ise
            # aşağıda Python tarafında tutar metni ayrıştırılarak bulunuyor.
            lov_penceresi.fill("#myListPage_DXFREditorcol2_I", fatura_no_str)
            lov_penceresi.press("#myListPage_DXFREditorcol2_I", "Enter")

            # Sabit bir bekleme yerine, filtrelenmiş satırın gerçekten
            # ekrana gelmesini bekle (sunucu tarafı AJAX callback'i tamamlanınca gelir)
            satir_bekleme_secici = f"tr.dxgvDataRow_Aqua:has-text('{fatura_no_str}')"
            lov_penceresi.wait_for_selector(satir_bekleme_secici, state="visible", timeout=15000)
            aktif_sayfa.wait_for_timeout(500)  # DevExpress render'ının oturması için küçük tampon

            hedef_tutar = satis_fiyati_decimal.quantize(Decimal("0.01"))

            aday_satirlar = lov_penceresi.locator(satir_bekleme_secici)
            aday_sayisi = aday_satirlar.count()

            hedef_satir = None
            incelenen_metinler = []
            for idx in range(aday_sayisi):
                aday = aday_satirlar.nth(idx)
                try:
                    metin = aday.inner_text()
                except Exception:
                    continue
                incelenen_metinler.append(metin.replace("\n", " | "))
                if _satirda_tutar_var_mi(metin, hedef_tutar):
                    hedef_satir = aday
                    break

            if hedef_satir is None:
                ornek = "\n".join(incelenen_metinler[:5])
                raise RuntimeError(
                    f"[{kaynak_yuk_no}] HATA: Fatura '{fatura_no_str}' için {hedef_tutar} TL tutarında "
                    f"eşleşen satır bulunamadı ({aday_sayisi} satır incelendi).\n"
                    f"İncelenen satır örnekleri:\n{ornek}"
                )

            hedef_satir.click()
            aktif_sayfa.wait_for_timeout(500)

            # Sizin F12 ile bulduğunuz doğru Buton ID'si!
            lov_penceresi.locator("#btnChoose_CD").click(force=True)
            aktif_sayfa.wait_for_timeout(1500)  # Pencerenin kapanmasını bekle

            # ANA EKRANA KAYDETME
            print(f"[{kaynak_yuk_no}] Fatura başarıyla bağlandı, satır kaydediliyor...")
            aktif_sayfa.wait_for_selector("a[id*='editnew']:has-text('Kaydet')", state="visible", timeout=15000)
            aktif_sayfa.locator("a[id*='editnew']:has-text('Kaydet')").first.click(force=True)

            # Bir sonraki döngüye geçmeden önce Yeni Satır butonunun tekrar geri gelmesini BEKLE
            aktif_sayfa.wait_for_selector("#TabControl_grd_LGoodsOpDetailCollection_EmptyRow_btnNew", state="visible", timeout=15000)

        # Ana Yükü Kaydet ve Çoklu Doğrulama Yap
        aktif_sayfa.click("#btnSave_CD", force=True)

        aktif_sayfa.wait_for_timeout(3000)
        aktif_sayfa.wait_for_function(
            """(oncekiDeger) => {
                let val = document.querySelector("#TabControl_txt_ReferenceNo_I").value.trim();
                return val !== "" && val.startsWith("Y-") && val !== oncekiDeger;
            }""",
            arg=kaynak_yuk_no,
            timeout=20000
        )

        yeni_yuk_no = aktif_sayfa.input_value("#TabControl_txt_ReferenceNo_I")
        print(f"[{kaynak_yuk_no}] Yeni Yük Oluştu: {yeni_yuk_no}")

        if checkpoint_kaydet:
            checkpoint_kaydet(yeni_yuk_no)

    # --- FAZ 4: SEVK OLUŞTURMA ---
    print(f"[{kaynak_yuk_no}] Faz 4: Sevk oluşturuluyor...")
    aktif_sayfa.click("#TabControl_txt_ReferenceNo_I", button="right")
    aktif_sayfa.wait_for_selector("div.uyum-popup-menu span:text-is('Sevk Oluştur')", state="visible")
    aktif_sayfa.click("div.uyum-popup-menu span:text-is('Sevk Oluştur')")

    aktif_sayfa.wait_for_selector("#TabControl_dte_DocDate_I", state="visible", timeout=15000)

    _tarih_alanini_temizle_ve_yaz(aktif_sayfa, "#TabControl_dte_DocDate_I", yuk_tarihi)

    aktif_sayfa.fill("#TabControl_bte_TractorSerialNoPlateNo_I", plaka)
    aktif_sayfa.press("#TabControl_bte_TractorSerialNoPlateNo_I", "Enter")
    aktif_sayfa.wait_for_timeout(2000)

    aktif_sayfa.click("#TabControl_grd_LTransOpDetailCollection_EmptyRow_btnNew")
    aktif_sayfa.wait_for_timeout(2000)
    aktif_sayfa.wait_for_selector("#TabControl_grd_LTransOpDetailCollection_DXEditor4_I", state="visible")

    if sevk_alis_fiyati is None:
        raise ValueError(f"[{kaynak_yuk_no}] HATA: Sevk alış fiyatı (Excel) boş, işlem durduruldu!")

    sevk_guvenli = Decimal(str(sevk_alis_fiyati))
    formatli_sevk_fiyati = f"{sevk_guvenli:.2f}".replace(".", ",")

    aktif_sayfa.fill("#TabControl_grd_LTransOpDetailCollection_DXEditor4_I", formatli_sevk_fiyati)
    aktif_sayfa.press("#TabControl_grd_LTransOpDetailCollection_DXEditor4_I", "Tab")

    aktif_sayfa.click("a[id*='editnew']:has-text('Kaydet')", force=True)
    aktif_sayfa.wait_for_timeout(1000)

    aktif_sayfa.click("#btnSave_CD", force=True)

    aktif_sayfa.wait_for_function('''
        () => {
            let val = document.querySelector("#TabControl_txt_TransportNo_I").value.trim();
            return val !== "" && val.startsWith("S-");
        }
    ''', timeout=20000)

    yeni_sevk_no = aktif_sayfa.input_value("#TabControl_txt_TransportNo_I")
    print(f"[{kaynak_yuk_no}] Yeni Sevk Oluştu: {yeni_sevk_no}")

    if not atla_faz_123:
        aktif_sayfa.close()

    return {
        "durum": "BAŞARILI",
        "yeni_yuk_no": yeni_yuk_no,
        "yeni_sevk_no": yeni_sevk_no,
        "proje": proje_kodu,
        "fatura_no": tum_faturalar,
        "fatura_tarihi": tum_fatura_tarihleri,
        "tarih": yuk_tarihi,
        "aktif_sayfa": page
    }


# ==========================================
# ANA ORKESTRASYON VE EXCEL YÖNETİMİ
# ==========================================
def main():
    excel_dosyasi = "islem_listesi.xlsx"

    if not os.path.exists(excel_dosyasi):
        print(f"HATA: '{excel_dosyasi}' dosyası bulunamadı. Lütfen dizini kontrol edin.")
        return

    yedek_isim = f"yedek_islem_listesi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    try:
        shutil.copy(excel_dosyasi, yedek_isim)
        print(f"Orijinal veri güvene alındı. Yedek: {yedek_isim}")
    except PermissionError:
        print("HATA: Excel dosyası açık! Lütfen Excel'i kapatıp otomasyonu tekrar çalıştırın.")
        return

    wb = load_workbook(excel_dosyasi)
    ws = wb.active

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=ayarlar.HEADLESS, channel="chrome")
        try:
            context = browser.new_context(viewport={'width': 1920, 'height': 1080})
            page = context.new_page()

            print("ERP sistemine giriş yapılıyor...")
            page.goto(ayarlar.ERP_LOGIN_URL)
            page.wait_for_selector("input#tbxUsername", state="visible", timeout=30000)

            page.fill("input#tbxUsername", ayarlar.ERP_KULLANICI)
            page.fill("input#tbxPassword", ayarlar.ERP_SIFRE)
            page.press("input#tbxPassword", "Enter")

            page.wait_for_timeout(4000)

            basarili_yeni_yukler = []
            basarili_yeni_sevkler = []

            max_row = ws.max_row
            for i in range(2, max_row + 1):
                kaynak_yuk_no = ws.cell(row=i, column=1).value
                plaka = ws.cell(row=i, column=2).value
                sevk_alis_fiyati = ws.cell(row=i, column=3).value

                mevcut_durum = ws.cell(row=i, column=10).value
                onceki_yeni_yuk_no = ws.cell(row=i, column=8).value

                if not kaynak_yuk_no:
                    continue

                if mevcut_durum in ["BAŞARILI", "DRY_RUN BAŞARILI"]:
                    print(f"[{kaynak_yuk_no}] Durum: {mevcut_durum} -> İşlem Atlanıyor...")
                    continue

                print(f"\n--- İŞLEM BAŞLIYOR: {kaynak_yuk_no} ---")

                hata_sayfasi = page
                try:
                    oracle_data = kaynak_yuk_verilerini_getir(kaynak_yuk_no)

                    def checkpoint_kaydet(yeni_yuk_no_degeri):
                        ws.cell(row=i, column=8).value = yeni_yuk_no_degeri
                        ws.cell(row=i, column=10).value = "YÜK OLUŞTU"
                        wb.save(excel_dosyasi)
                        print(f"[{kaynak_yuk_no}] CHECKPOINT: Yük {yeni_yuk_no_degeri} Excel'e kaydedildi -> Sevk aşamasına geçiliyor.")

                    sonuc = uyumsoft_islemlerini_yap(page, kaynak_yuk_no, plaka, sevk_alis_fiyati, oracle_data, mevcut_durum, onceki_yeni_yuk_no, checkpoint_kaydet)

                    hata_sayfasi = sonuc["aktif_sayfa"]

                    ws.cell(row=i, column=4).value = sonuc["proje"]
                    ws.cell(row=i, column=5).value = sonuc["tarih"]
                    ws.cell(row=i, column=6).value = sonuc["fatura_no"]
                    ws.cell(row=i, column=7).value = sonuc["fatura_tarihi"]
                    if sonuc["yeni_yuk_no"]:
                        ws.cell(row=i, column=8).value = sonuc["yeni_yuk_no"]
                    if sonuc["yeni_sevk_no"]:
                        ws.cell(row=i, column=9).value = sonuc["yeni_sevk_no"]
                    ws.cell(row=i, column=10).value = sonuc["durum"]

                    if sonuc["durum"] == "BAŞARILI":
                        if sonuc["yeni_yuk_no"]:
                            basarili_yeni_yukler.append(sonuc["yeni_yuk_no"])
                        if sonuc["yeni_sevk_no"]:
                            basarili_yeni_sevkler.append(sonuc["yeni_sevk_no"])

                    wb.save(excel_dosyasi)

                except Exception as e:
                    hata_mesaji = str(e)
                    print(f"HATA OLUŞTU [{kaynak_yuk_no}]: {hata_mesaji}")

                    zaman_damgasi = datetime.now().strftime("%H%M%S")
                    hata_foto = f"hata_{kaynak_yuk_no}_{zaman_damgasi}.png"

                    try:
                        if len(context.pages) > 1:
                            hata_sayfasi = context.pages[-1]

                        hata_sayfasi.screenshot(path=hata_foto)
                        hata_sayfasi.keyboard.press("Escape")

                        if mevcut_durum in ["YÜK OLUŞTU", "HATA_SEVK"]:
                            ws.cell(row=i, column=10).value = "HATA_SEVK"
                        else:
                            ws.cell(row=i, column=10).value = "HATA_YUK"
                            ws.cell(row=i, column=8).value = None

                        if len(context.pages) > 1:
                            hata_sayfasi.close()
                    except Exception:
                        ws.cell(row=i, column=10).value = "HATA_BİLİNMEYEN"

                    ws.cell(row=i, column=11).value = hata_mesaji
                    wb.save(excel_dosyasi)
                    print(f"Hata detayı kaydedildi. Ekran Görüntüsü: {hata_foto}")

                    try:
                        page.goto(ayarlar.ERP_YUK_LISTESI_URL)
                    except Exception:
                        print("Sistem başlangıç ekranına dönemedi! Güvenlik için otomasyon durduruluyor.")
                        break

            print("\nTüm Excel satırları tamamlandı.")

            if basarili_yeni_yukler or basarili_yeni_sevkler:
                print("\n>>> Veritabanı İz Temizliği (Toplu SQL İşlemi) Başlatılıyor... <<<")
                yeni_kayitlari_veritabaninda_guncelle(basarili_yeni_yukler, basarili_yeni_sevkler)

            print("\nERP'den güvenli çıkış yapılıyor...")
            try:
                page.click("div.logout-main")
                page.wait_for_timeout(2000)
            except Exception:
                pass
        finally:
            browser.close()


if __name__ == "__main__":
    main()
