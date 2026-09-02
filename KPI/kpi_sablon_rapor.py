"""
Referans KPI şablonunu doldurur: VERİ + Filo Detay → pivot sayfaları Excel'de güncellenir.

Pivotlu şablonlar openpyxl ile kaydedildiğinde bozulabildiği için (Excel açamaz),
Windows'ta veri yazımı doğrudan Excel COM ile yapılır.

Kullanım:
  1. Temmuz KPI dosyanızı KPI/referans/kpi_sablon.xlsx olarak kaydedin
  2. python kpi_rapor_olustur.py
"""
from __future__ import annotations

import re
import shutil
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import ayarlar
from kpi_kiralk_arac import FILO_DETAY_SUTUNLARI, kiralk_arac_detay_getir, kiralk_arac_semasi_hazir
from kpi_veri import hucre_degeri, veri_satirlari_getir, veri_semasi_hazir
from oracle_baglanti import baglanti_yonet

_KPI_KOKU = Path(__file__).resolve().parent


def _progress(mesaj: str) -> None:
    """Konsola anında yazar (Windows cmd tamponlaması / sessiz bekleme önlenir)."""
    print(mesaj, flush=True)

# Şablon başlığı (normalize) → Oracle kolon adı — Temmuz KPI şablonu için varsayılanlar
_SABLON_KOLON_VARSAYILAN: dict[str, str] = {
    "YUK_NO": "YUK_NO",
    "YUK_NUMARASI": "YUK_NO",
    "SEVK_NO": "SEVK_NO",
    "SEVK_NUMARASI": "SEVK_NO",
    "YUK_TARIHI": "YUK_TARIHI",
    "SEVK_TARIHI": "SEVK_TARIHI",
    "PROJE_KODU": "PROJE_KODU",
    "PROJE": "PROJE_KODU",
    "PLAKA": "PLAKA",
    "ARAC_TIPI": "ARAC_TIPI",
    "MUSTERI_KODU": "MUSTERI_KODU",
    "MUSTERI_ADI": "MUSTERI_ADI",
    "SUBE": "SUBE",
    "SUBE_KODU": "SUBE_KODU",
    "SUBE_ADI": "SUBE",
    "SATIS_TUTAR": "SATIS_TUTAR",
    "SATIS": "SATIS_TUTAR",
    "NET_SATIS": "SATIS_TUTAR",
    "TOPLAM_SATIS": "TOPLAM_SATIS",
    "ALIS_TUTAR": "ALIS_TUTAR",
    "ALIS": "ALIS_TUTAR",
    "TOPLAM_ALIS": "TOPLAM_ALIS",
    "KAR_ZARAR": "KAR_ZARAR",
    "NET_KAR_ZARAR": "NET_KAR_ZARAR",
    "KAR_ZARAR_TUTAR": "KAR_ZARAR",
    "MARJ_YUZDE": "MARJ_YUZDE",
    "MARJ_ORANI": "MARJ_ORANI",
    "MARJ": "MARJ_YUZDE",
    "YUK_FIYAT_TIP_KODU": "YUK_FIYAT_TIP_KODU",
    "MAL_TIPI": "YUK_FIYAT_TIP_KODU",
    "MAL_TIP": "YUK_FIYAT_TIP_KODU",
    "MAL_TIP_KODU": "YUK_FIYAT_TIP_KODU",
    "ALIS_TUTARI": "ALIS_TUTAR",
    "SATIS_TUTARI": "SATIS_TUTAR",
    "ALIS_IADE_TUTARI": "ALIS_IADE_TUTAR",
    "SATIS_IADE_TUTARI": "SATIS_IADE_TUTAR",
    "TOPLAM_KAR_ZARAR": "KAR_ZARAR",
    "TOPLAM_KAR_ZARAR_TUTAR": "KAR_ZARAR",
}


def _referans_klasoru() -> Path:
    return _KPI_KOKU / "referans"


def sablon_yolu() -> Path:
    dosya = getattr(ayarlar, "KPI_SABLON_DOSYASI", "kpi_sablon.xlsx")
    yol = Path(dosya)
    if yol.is_absolute():
        return yol
    return _referans_klasoru() / dosya


def _raporlar_klasoru() -> Path:
    klasor = _KPI_KOKU / "raporlar"
    klasor.mkdir(exist_ok=True)
    return klasor


def _cikti_yolu(sablon: Path | None = None) -> Path:
    dosya = getattr(ayarlar, "KPI_RAPOR_DOSYASI", None)
    if dosya:
        yol = Path(dosya)
        if yol.is_absolute():
            return yol
        return _raporlar_klasoru() / dosya
    suffix = (sablon or sablon_yolu()).suffix or ".xlsx"
    return _raporlar_klasoru() / f"kpi_rapor{suffix}"


def _normalize_kolon(adi: str) -> str:
    if adi is None:
        return ""
    metin = str(adi).strip().upper()
    metin = unicodedata.normalize("NFKD", metin)
    metin = "".join(c for c in metin if not unicodedata.combining(c))
    metin = re.sub(r"[^A-Z0-9]+", "_", metin)
    return metin.strip("_")


def _sayfa_bul(wb, adlar: list[str]) -> Worksheet | None:
    ad_norm = {a.strip().upper() for a in adlar}
    for sheet in wb.worksheets:
        if sheet.title.strip().upper() in ad_norm:
            return sheet
    for ad in adlar:
        if ad in wb.sheetnames:
            return wb[ad]
    return None


def _basliklari_oku(ws: Worksheet, satir: int = 1) -> list[str | None]:
    max_col = ws.max_column or 1
    return [ws.cell(row=satir, column=c).value for c in range(1, max_col + 1)]


def _kolon_esleme(
    basliklar: list[str | None],
    veri_anahtarlari: list[str],
    tablo_sol: int = 1,
) -> dict[int, str]:
    normalized_veri = {_normalize_kolon(k): k for k in veri_anahtarlari}
    for sablon_norm, oracle in _SABLON_KOLON_VARSAYILAN.items():
        if oracle in veri_anahtarlari:
            normalized_veri.setdefault(sablon_norm, oracle)
    ek = getattr(ayarlar, "KPI_KOLON_ESLEME", {}) or {}
    for sablon, oracle in ek.items():
        normalized_veri[_normalize_kolon(sablon)] = oracle

    esleme: dict[int, str] = {}
    for i, baslik in enumerate(basliklar):
        if baslik is None or str(baslik).strip() == "":
            continue
        norm = _normalize_kolon(str(baslik))
        if norm in normalized_veri:
            esleme[tablo_sol + i] = normalized_veri[norm]
    return esleme


def _eslesmeyen_basliklar(
    basliklar: list[Any],
    esleme: dict[int, str],
    tablo_sol: int = 1,
) -> list[str]:
    sonuc: list[str] = []
    for i, baslik in enumerate(basliklar):
        if baslik is None or str(baslik).strip() == "":
            continue
        col_idx = tablo_sol + i
        if col_idx not in esleme:
            sonuc.append(str(baslik).strip())
    return sonuc


def _sutun1_arteakt_mi(baslik: Any) -> bool:
    if baslik is None:
        return False
    return _normalize_kolon(str(baslik)) in ("SUTUN1", "COLUMN1")


def _sayfayi_temizle_yaz(
    ws: Worksheet,
    baslik_satiri: int,
    satirlar: list[dict[str, Any]],
    sabit_kolonlar: list[str] | None = None,
):
    basliklar = _basliklari_oku(ws, baslik_satiri)
    if not satirlar:
        max_row = ws.max_row or baslik_satiri
        for row in range(baslik_satiri + 1, max_row + 1):
            for col in range(1, len(basliklar) + 1):
                ws.cell(row=row, column=col, value=None)
        return 0

    anahtarlar = sabit_kolonlar or list(satirlar[0].keys())
    esleme = _kolon_esleme(basliklar, anahtarlar)

    if not esleme and sabit_kolonlar:
        esleme = {i + 1: k for i, k in enumerate(sabit_kolonlar) if i < len(basliklar)}

    max_row = ws.max_row or baslik_satiri
    for row in range(baslik_satiri + 1, max(max_row, baslik_satiri + len(satirlar)) + 1):
        for col in range(1, len(basliklar) + 1):
            ws.cell(row=row, column=col, value=None)

    for i, satir in enumerate(satirlar):
        row_no = baslik_satiri + 1 + i
        if sabit_kolonlar:
            for col_idx, kolon in enumerate(sabit_kolonlar, 1):
                ws.cell(row=row_no, column=col_idx, value=hucre_degeri(satir.get(kolon)))
        else:
            for col_idx, kolon in esleme.items():
                ws.cell(row=row_no, column=col_idx, value=hucre_degeri(satir.get(kolon)))

    return len(satirlar)


def _hucre_gorunen_uzunluk(deger: Any) -> int:
    if deger is None:
        return 0
    if isinstance(deger, bool):
        return 4
    if isinstance(deger, (datetime, date)):
        return len(deger.strftime("%d.%m.%Y"))
    if isinstance(deger, float):
        metin = f"{deger:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return len(metin) + 2
    if isinstance(deger, int):
        metin = f"{deger:,}".replace(",", ".")
        return len(metin)
    return len(str(deger))


def _sayfa_sutunlarini_genislet(
    ws: Worksheet,
    baslik_satiri: int = 1,
    veri_satir_sayisi: int = 0,
    min_genislik: float = 10,
    max_genislik: float = 55,
    padding: float = 2,
):
    """Doldurulan sayfada sütun genişliklerini içeriğe göre ayarlar (####### önler)."""
    son_satir = baslik_satiri + max(veri_satir_sayisi, 0)
    if son_satir < baslik_satiri:
        return

    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        en_uzun = 0
        for row in range(baslik_satiri, son_satir + 1):
            deger = ws.cell(row=row, column=col).value
            en_uzun = max(en_uzun, _hucre_gorunen_uzunluk(deger))
        if en_uzun <= 0:
            continue
        harf = get_column_letter(col)
        mevcut = ws.column_dimensions[harf].width or min_genislik
        yeni = min(max(en_uzun + padding, min_genislik), max_genislik)
        ws.column_dimensions[harf].width = max(mevcut, yeni)


def _excel_kullanilabilir() -> bool:
    if getattr(ayarlar, "KPI_EXCEL_KULLAN", True) is False:
        return False
    try:
        import win32com.client  # type: ignore  # noqa: F401
    except ImportError:
        return False
    return True


def _com_satir_oku(deger: Any) -> list[Any]:
    if deger is None:
        return []
    if not isinstance(deger, tuple):
        return [deger]
    if not deger:
        return []
    if not isinstance(deger[0], tuple):
        return list(deger)
    return list(deger[0])


def _com_basliklari_oku(
    sheet,
    baslik_satiri: int,
    tablo_sol: int = 1,
    tablo_sag: int | None = None,
) -> list[Any]:
    """Başlık satırını fiziksel hücrelerden okur (ListColumns kullanılmaz — kaydırma hatası önlenir)."""
    if tablo_sag is None:
        tablo_sag = tablo_sol + 59
        try:
            used = sheet.UsedRange
            tablo_sag = max(int(used.Column) + int(used.Columns.Count) - 1, tablo_sag)
        except Exception:
            pass
    ham = sheet.Range(
        sheet.Cells(baslik_satiri, tablo_sol),
        sheet.Cells(baslik_satiri, tablo_sag),
    ).Value
    basliklar = _com_satir_oku(ham)
    while basliklar and basliklar[-1] is None:
        basliklar.pop()
    return basliklar


def _com_sutun1_baslangic_temizle(sheet, baslik_satiri: int) -> bool:
    """Resize hatasıyla A sütununa eklenen 'Sütun1' artefaktını kaldırır."""
    try:
        a1 = sheet.Cells(baslik_satiri, 1).Value
        b1 = sheet.Cells(baslik_satiri, 2).Value
    except Exception:
        return False
    if not _sutun1_arteakt_mi(a1):
        return False
    if b1 is None or str(b1).strip() == "":
        return False
    try:
        sheet.Columns(1).Delete()
        return True
    except Exception:
        return False


def _com_listobject_bul(sheet, baslik_satiri: int):
    """Sayfadaki Excel Tablosunu (ListObject) bulur — pivot kaynağı genelde budur."""
    try:
        adet = int(sheet.ListObjects.Count)
    except Exception:
        return None
    for i in range(1, adet + 1):
        lo = sheet.ListObjects(i)
        try:
            if int(lo.HeaderRowRange.Row) == baslik_satiri:
                return lo
        except Exception:
            continue
    if adet >= 1:
        return sheet.ListObjects(1)
    return None


def _com_tablo_konumu(sheet, baslik_satiri: int) -> tuple[int, int, Any | None]:
    """Excel tablosunun gerçek sol sütun ve kolon sayısını döndürür."""
    lo = _com_listobject_bul(sheet, baslik_satiri)
    if lo is not None:
        try:
            hdr = lo.HeaderRowRange
            tablo_sol = int(hdr.Column)
            kolon_sayisi = int(hdr.Columns.Count)
            if kolon_sayisi > 0:
                return tablo_sol, kolon_sayisi, lo
        except Exception:
            pass
    basliklar = _com_basliklari_oku(sheet, baslik_satiri, 1)
    return 1, max(len(basliklar), 1), lo


def _com_veri_matrisi_hazirla(
    basliklar: list[Any],
    satirlar: list[dict[str, Any]],
    sabit_kolonlar: list[str] | None,
    kolon_sayisi: int,
    tablo_sol: int = 1,
) -> list[tuple[Any, ...]]:
    if not satirlar:
        return []

    anahtarlar = sabit_kolonlar or list(satirlar[0].keys())
    esleme = _kolon_esleme(basliklar, anahtarlar, tablo_sol)
    if not esleme and sabit_kolonlar:
        esleme = {
            tablo_sol + i: k for i, k in enumerate(sabit_kolonlar) if i < len(basliklar)
        }

    matris: list[tuple[Any, ...]] = []
    tablo_sag = tablo_sol + kolon_sayisi - 1
    for satir in satirlar:
        if sabit_kolonlar:
            satir_verisi: list[Any] = [
                hucre_degeri(satir.get(k)) for k in sabit_kolonlar[:kolon_sayisi]
            ]
            if len(satir_verisi) < kolon_sayisi:
                satir_verisi.extend([None] * (kolon_sayisi - len(satir_verisi)))
        else:
            satir_verisi = [None] * kolon_sayisi
            for col_idx, kolon in esleme.items():
                if tablo_sol <= col_idx <= tablo_sag:
                    satir_verisi[col_idx - tablo_sol] = hucre_degeri(satir.get(kolon))
        matris.append(tuple(satir_verisi))
    return matris


def _com_araliga_yaz(hedef, matris: list[tuple[Any, ...]]):
    if not matris:
        return
    if len(matris) == 1:
        hedef.Value = matris[0]
    else:
        hedef.Value = tuple(matris)


def _com_sayfaya_yaz(
    sheet,
    baslik_satiri: int,
    satirlar: list[dict[str, Any]],
    sabit_kolonlar: list[str] | None = None,
) -> tuple[int, dict[int, str], list[str], int, int]:
    """Pivot kaynağına yazar — tablo sütun sınırları korunur, A'ya genişletme yapılmaz."""
    if _com_sutun1_baslangic_temizle(sheet, baslik_satiri):
        lo = _com_listobject_bul(sheet, baslik_satiri)
        if lo is not None:
            try:
                lo = sheet.ListObjects(lo.Name)
            except Exception:
                pass

    tablo_sol, kolon_sayisi, lo = _com_tablo_konumu(sheet, baslik_satiri)
    tablo_sag = tablo_sol + kolon_sayisi - 1
    basliklar = _com_basliklari_oku(sheet, baslik_satiri, tablo_sol, tablo_sag)
    if basliklar:
        kolon_sayisi = len(basliklar)
        tablo_sag = tablo_sol + kolon_sayisi - 1

    anahtarlar = sabit_kolonlar or (list(satirlar[0].keys()) if satirlar else [])
    esleme = _kolon_esleme(basliklar, anahtarlar, tablo_sol)
    if not esleme and sabit_kolonlar:
        esleme = {
            tablo_sol + i: k for i, k in enumerate(sabit_kolonlar) if i < len(basliklar)
        }
    eslesmeyen = (
        _eslesmeyen_basliklar(basliklar, esleme, tablo_sol) if not sabit_kolonlar else []
    )

    if not satirlar:
        return 0, esleme, eslesmeyen, kolon_sayisi, tablo_sol

    matris = _com_veri_matrisi_hazirla(
        basliklar, satirlar, sabit_kolonlar, kolon_sayisi, tablo_sol
    )
    if not matris:
        return 0, esleme, eslesmeyen, kolon_sayisi, tablo_sol

    son_satir = baslik_satiri + len(matris)
    hedef = sheet.Range(
        sheet.Cells(baslik_satiri + 1, tablo_sol),
        sheet.Cells(son_satir, tablo_sag),
    )

    if lo is not None:
        try:
            hdr = lo.HeaderRowRange
            resize_sol = int(hdr.Column)
            resize_sag = resize_sol + int(hdr.Columns.Count) - 1
            yeni_tablo = sheet.Range(
                sheet.Cells(baslik_satiri, resize_sol),
                sheet.Cells(son_satir, resize_sag),
            )
            lo.Resize(yeni_tablo)
        except Exception:
            pass

    _com_araliga_yaz(hedef, matris)

    return len(satirlar), esleme, eslesmeyen, kolon_sayisi, tablo_sol


def _com_pivot_kaynak_guncelle(
    wb,
    sayfa_adi: str,
    baslik_satiri: int,
    satir_sayisi: int,
    kolon_sayisi: int,
    tablo_sol: int = 1,
):
    if satir_sayisi <= 0:
        return
    son_satir = baslik_satiri + satir_sayisi
    bas_kolon = get_column_letter(max(tablo_sol, 1))
    son_kolon = get_column_letter(max(tablo_sol + kolon_sayisi - 1, tablo_sol))
    yeni_kaynak = f"'{sayfa_adi}'!${bas_kolon}${baslik_satiri}:${son_kolon}${son_satir}"
    try:
        for i in range(1, int(wb.PivotCaches().Count) + 1):
            pc = wb.PivotCaches(i)
            src = str(pc.SourceData or "")
            if sayfa_adi.upper() not in src.upper():
                continue
            # Tablo adına bağlı pivot (örn. 'VERİ'!Tablo1) — ListObject.Resize yeterli
            if "$" not in src:
                continue
            pc.SourceData = yeni_kaynak
    except Exception:
        pass


def _excel_uygulama_ac():
    import win32com.client  # type: ignore

    try:
        excel = win32com.client.gencache.EnsureDispatch("Excel.Application")
    except Exception:
        excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    excel.EnableEvents = False
    return excel


def _autofit_ayarlari() -> tuple[float, float, float, int, dict[str, float]]:
    ozel = getattr(ayarlar, "KPI_SUTUN_GENISLIK", None) or {}
    if not isinstance(ozel, dict):
        ozel = {}
    return (
        float(getattr(ayarlar, "KPI_SUTUN_MIN_GENISLIK", 10)),
        float(getattr(ayarlar, "KPI_PARA_SUTUN_MIN_GENISLIK", 24)),
        float(getattr(ayarlar, "KPI_SUTUN_MAX_GENISLIK", 55)),
        int(getattr(ayarlar, "KPI_AUTOFIT_MAX_SATIR", 400)),
        {str(k).upper(): float(v) for k, v in ozel.items()},
    )


def _com_sutun_genisligi_ayarla(sheet, col: int | str, genislik: float, max_w: float) -> None:
    try:
        mevcut = float(sheet.Columns(col).ColumnWidth)
        hedef = min(max(genislik, mevcut), max_w)
        if hedef > mevcut:
            sheet.Columns(col).ColumnWidth = hedef
    except Exception:
        pass


def _com_sutunlari_genislet(sheet, satir_limit: int | None = None) -> None:
    """UsedRange AutoFit + #### tespiti — pivot tutar sütunları (Alış/Satış) için."""
    min_w, para_w, max_w, varsayilan_satir_limit, ozel_genislik = _autofit_ayarlari()
    if satir_limit is None:
        satir_limit = varsayilan_satir_limit

    try:
        used = sheet.UsedRange
        if used is None:
            return
    except Exception:
        return

    try:
        used.Columns.AutoFit()
    except Exception:
        pass

    try:
        first_col = int(used.Column)
        col_count = int(used.Columns.Count)
        first_row = int(used.Row)
        row_count = min(int(used.Rows.Count), satir_limit)
    except Exception:
        return

    for offset in range(col_count):
        col = first_col + offset
        hedef = min_w
        kesin_para = False
        for row in range(first_row, first_row + row_count):
            try:
                text = str(sheet.Cells(row, col).Text or "").strip()
            except Exception:
                continue
            if not text:
                continue
            genislik_ihtiyaci = min(len(text) * 1.12 + 2.0, max_w)
            if "#" in text:
                hedef = max(hedef, genislik_ihtiyaci, para_w)
                kesin_para = True
                break
            if any(ch in text for ch in "₺%") or ("." in text and "," in text) or (
                "," in text and any(c.isdigit() for c in text)
            ):
                hedef = max(hedef, genislik_ihtiyaci, para_w)
                kesin_para = True
            else:
                hedef = max(hedef, genislik_ihtiyaci)

        if kesin_para:
            hedef = max(hedef, para_w)

        _com_sutun_genisligi_ayarla(sheet, col, hedef, max_w)

    # Özet vb.: Alış sütunu (C) — şablonda tutar genelde burada
    alis_harf = str(getattr(ayarlar, "KPI_ALIS_SUTUN_HARFI", "C")).upper()
    alis_genislik = float(getattr(ayarlar, "KPI_ALIS_SUTUN_GENISLIK", para_w + 4))
    _com_sutun_genisligi_ayarla(sheet, alis_harf, alis_genislik, max_w)

    for harf, gen in ozel_genislik.items():
        _com_sutun_genisligi_ayarla(sheet, harf, gen, max_w)


def _excel_hesapla(excel) -> None:
    """Pivot sonrası hesaplama — hızlı modda tam yeniden derleme yapılmaz."""
    hizli = getattr(ayarlar, "KPI_HIZLI_MOD", True)
    try:
        excel.CalculateUntilAsyncQueriesDone()
    except Exception:
        pass
    try:
        if hizli:
            excel.Calculate()
        else:
            excel.CalculateFullRebuild()
    except Exception:
        pass


def _excel_sayfa_bul(wb, adlar: list[str]):
    ad_norm = {a.strip().upper() for a in adlar}
    for sheet in wb.Worksheets:
        if str(sheet.Name).strip().upper() in ad_norm:
            return sheet
    return None


def _excel_sablon_doldur(
    dosya_yolu: Path,
    veri_satirlari: list[dict[str, Any]],
    filo_satirlari: list[dict[str, Any]],
    veri_sayfa_adlari: list[str],
    filo_sayfa_adlari: list[str],
    veri_baslik_satiri: int,
    filo_baslik_satiri: int,
    pivot_yenile: bool,
    sutun_autofit: bool,
) -> tuple[bool, str | None, int, int, list[str]]:
    """Şablon kopyasına Excel COM ile veri yazar, pivot yeniler, sütunları genişletir."""
    excel = None
    wb = None
    uyarilar: list[str] = []
    eslesmeyen_veri: list[str] = []
    dosya = str(dosya_yolu.resolve())

    try:
        _progress("  [Excel] Uygulama açılıyor...")
        excel = _excel_uygulama_ac()
        _progress(f"  [Excel] Şablon açılıyor: {dosya_yolu.name}")
        wb = excel.Workbooks.Open(
            Filename=dosya,
            UpdateLinks=0,
            ReadOnly=False,
            Notify=False,
        )

        ws_veri = _excel_sayfa_bul(wb, veri_sayfa_adlari)
        if ws_veri is None:
            return False, f"VERİ sayfası bulunamadı: {veri_sayfa_adlari}", 0, 0, []

        ws_filo = _excel_sayfa_bul(wb, filo_sayfa_adlari)
        if ws_filo is None:
            return False, f"Filo Detay sayfası bulunamadı: {filo_sayfa_adlari}", 0, 0, []

        _progress(f"  [Excel] VERİ yazılıyor ({len(veri_satirlari)} satır)...")
        veri_adet, _, eslesmeyen_veri, veri_kolon, veri_tablo_sol = _com_sayfaya_yaz(
            ws_veri, veri_baslik_satiri, veri_satirlari
        )
        _com_pivot_kaynak_guncelle(
            wb, ws_veri.Name, veri_baslik_satiri, veri_adet, veri_kolon, veri_tablo_sol
        )

        _progress(f"  [Excel] Filo Detay yazılıyor ({len(filo_satirlari)} satır)...")
        filo_yaz = [{k: r.get(k) for k in FILO_DETAY_SUTUNLARI} for r in filo_satirlari]
        filo_adet, _, _, filo_kolon, filo_tablo_sol = _com_sayfaya_yaz(
            ws_filo, filo_baslik_satiri, filo_yaz, sabit_kolonlar=FILO_DETAY_SUTUNLARI
        )
        _com_pivot_kaynak_guncelle(
            wb, ws_filo.Name, filo_baslik_satiri, filo_adet, filo_kolon, filo_tablo_sol
        )

        if pivot_yenile:
            _progress("  [Excel] Pivotlar yenileniyor...")
            try:
                wb.RefreshAll()
            except Exception as exc:
                uyarilar.append(f"pivot yenileme: {exc}")
            try:
                _excel_hesapla(excel)
            except Exception as exc:
                uyarilar.append(f"hesaplama: {exc}")

        if sutun_autofit:
            _progress("  [Excel] Özet sayfaları genişletiliyor...")
            for sheet in wb.Worksheets:
                if int(sheet.Visible) != -1:
                    continue
                if sheet.Name in (ws_veri.Name, ws_filo.Name):
                    continue
                try:
                    _com_sutunlari_genislet(sheet)
                except Exception as exc:
                    uyarilar.append(f"{sheet.Name} AutoFit: {exc}")

        _progress("  [Excel] Kaydediliyor...")
        wb.Save()
        wb.Close(SaveChanges=True)
        wb = None

        mesaj = f"kısmi uyarı: {'; '.join(uyarilar)}" if uyarilar else None
        return True, mesaj, veri_adet, filo_adet, eslesmeyen_veri

    except Exception as exc:
        return False, str(exc), 0, 0, []

    finally:
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


def _excel_islemleri(dosya_yolu: Path, pivot_yenile: bool = True) -> tuple[bool, str | None]:
    """Mevcut dosyada yalnızca pivot yenile + AutoFit (eski akış / test)."""
    try:
        import win32com.client  # type: ignore
    except ImportError:
        return False, "pywin32 kurulu değil"

    excel = None
    wb = None
    uyarilar: list[str] = []
    dosya = str(dosya_yolu.resolve())

    try:
        excel = _excel_uygulama_ac()
        wb = excel.Workbooks.Open(
            Filename=dosya,
            UpdateLinks=0,
            ReadOnly=False,
            Notify=False,
        )

        if pivot_yenile:
            try:
                wb.RefreshAll()
            except Exception as exc:
                uyarilar.append(f"pivot yenileme: {exc}")
            try:
                _excel_hesapla(excel)
            except Exception as exc:
                uyarilar.append(f"hesaplama: {exc}")

        autofit_sayisi = 0
        for sheet in wb.Worksheets:
            if int(sheet.Visible) != -1:
                continue
            try:
                _com_sutunlari_genislet(sheet)
                autofit_sayisi += 1
            except Exception as exc:
                uyarilar.append(f"{sheet.Name} AutoFit: {exc}")

        wb.Save()
        wb.Close(SaveChanges=True)
        wb = None

        if autofit_sayisi == 0:
            mesaj = "; ".join(uyarilar) if uyarilar else "Görünür sayfa bulunamadı"
            return False, mesaj

        if uyarilar:
            return True, f"kısmi uyarı: {'; '.join(uyarilar)}"
        return True, None

    except Exception as exc:
        return False, str(exc)

    finally:
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


def pivot_yenile(dosya_yolu: Path) -> tuple[bool, str | None]:
    """Windows + Excel kurulu ise pivotları yeniler ve sütunları genişletir."""
    return _excel_islemleri(dosya_yolu, pivot_yenile=True)


def _openpyxl_yedek_yaz(
    hedef: Path,
    veri_satirlari: list[dict[str, Any]],
    filo_satirlari: list[dict[str, Any]],
    veri_sayfa_adlari: list[str],
    filo_sayfa_adlari: list[str],
    veri_baslik_satiri: int,
    filo_baslik_satiri: int,
) -> tuple[int, int]:
    """Excel yoksa yedek — pivotlu şablon Excel'de açılamayabilir."""
    wb = load_workbook(hedef, keep_vba=True)

    ws_veri = _sayfa_bul(wb, veri_sayfa_adlari)
    if ws_veri is None:
        wb.close()
        raise ValueError(f"Şablonda VERİ sayfası bulunamadı: {veri_sayfa_adlari}")

    ws_filo = _sayfa_bul(wb, filo_sayfa_adlari)
    if ws_filo is None:
        wb.close()
        raise ValueError(f"Şablonda Filo Detay sayfası bulunamadı: {filo_sayfa_adlari}")

    veri_adet = _sayfayi_temizle_yaz(ws_veri, veri_baslik_satiri, veri_satirlari)
    filo_yaz = [{k: r.get(k) for k in FILO_DETAY_SUTUNLARI} for r in filo_satirlari]
    filo_adet = _sayfayi_temizle_yaz(
        ws_filo, filo_baslik_satiri, filo_yaz, sabit_kolonlar=FILO_DETAY_SUTUNLARI
    )
    _sayfa_sutunlarini_genislet(ws_veri, veri_baslik_satiri, veri_adet)
    _sayfa_sutunlarini_genislet(ws_filo, filo_baslik_satiri, filo_adet)
    wb.save(hedef)
    wb.close()
    return veri_adet, filo_adet


def _bind_olustur(bas: str, bit: str) -> dict:
    bind = {"bas": bas, "bit": bit}
    if getattr(ayarlar, "CO_CODE", None):
        bind["co_code"] = ayarlar.CO_CODE
    if getattr(ayarlar, "BRANCH_CODE", None):
        bind["branch_code"] = ayarlar.BRANCH_CODE
    return bind


def _tarih_araligi() -> tuple[str, str]:
    bugun = datetime.now().date()
    bas = getattr(ayarlar, "KPI_BASLANGIC_TARIHI", bugun.replace(day=1).strftime("%d.%m.%Y"))
    bit = getattr(ayarlar, "KPI_BITIS_TARIHI", bugun.strftime("%d.%m.%Y"))
    return bas, bit


def sablon_rapor_olustur(
    sablon: Path | None = None,
    cikti: Path | None = None,
    pivot_yenile_calistir: bool | None = None,
) -> str:
    kaynak = sablon or sablon_yolu()
    hedef = cikti or _cikti_yolu(kaynak)

    if not kaynak.exists():
        raise FileNotFoundError(
            f"KPI şablonu bulunamadı: {kaynak}\n"
            f"Temmuz KPI dosyanızı bu konuma 'kpi_sablon.xlsx' adıyla kopyalayın."
        )

    bas, bit = _tarih_araligi()
    bind = _bind_olustur(bas, bit)

    _progress(f"KPI raporu — dönem: {bas} — {bit}")
    _progress(f"  Şablon: {kaynak.name}")

    hedef.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(kaynak, hedef)
    _progress(f"  Çıktı kopyalandı: {hedef.name}")

    veri_satirlari: list[dict] = []
    filo_satirlari: list[dict] = []

    _progress("  Oracle bağlantısı açılıyor...")
    with baglanti_yonet() as baglanti:
        cursor = baglanti.cursor()
        veri_ok, veri_mesaj = veri_semasi_hazir()
        if veri_ok:
            _progress("  Oracle VERİ sorgusu çalışıyor (birkaç dakika sürebilir)...")
            veri_satirlari = veri_satirlari_getir(cursor, bas, bit, bind)
            _progress(f"  Oracle VERİ tamam: {len(veri_satirlari)} satır")
        else:
            _progress(f"  Uyarı: VERİ atlandı — {veri_mesaj}")
        filo_ok, filo_mesaj = kiralk_arac_semasi_hazir()
        if filo_ok:
            _progress("  Oracle Filo Detay sorgusu çalışıyor...")
            filo_satirlari = kiralk_arac_detay_getir(cursor, bas, bit, bind)
            _progress(f"  Oracle Filo tamam: {len(filo_satirlari)} satır")
        else:
            _progress(f"  Uyarı: Filo Detay atlandı — {filo_mesaj}")

    veri_sayfa_adlari = getattr(ayarlar, "KPI_VERI_SAYFA_ADLARI", ["VERİ", "VERI", "Veri"])
    filo_sayfa_adlari = getattr(ayarlar, "KPI_FILO_SAYFA_ADLARI", ["Filo Detay", "Filo detay"])
    veri_baslik_satiri = int(getattr(ayarlar, "KPI_VERI_BASLIK_SATIRI", 1))
    filo_baslik_satiri = int(getattr(ayarlar, "KPI_FILO_BASLIK_SATIRI", 1))

    if pivot_yenile_calistir is None:
        pivot_yenile_calistir = getattr(ayarlar, "KPI_PIVOT_YENILE", True)
    sutun_autofit = getattr(ayarlar, "KPI_SUTUN_AUTOFIT", True)

    excel_mesaj: str | None = None
    veri_adet = 0
    filo_adet = 0
    eslesmeyen_veri: list[str] = []

    if _excel_kullanilabilir():
        _progress("  Excel ile şablon dolduruluyor...")
        excel_ok, excel_mesaj, veri_adet, filo_adet, eslesmeyen_veri = _excel_sablon_doldur(
            hedef,
            veri_satirlari,
            filo_satirlari,
            veri_sayfa_adlari,
            filo_sayfa_adlari,
            veri_baslik_satiri,
            filo_baslik_satiri,
            pivot_yenile=pivot_yenile_calistir,
            sutun_autofit=sutun_autofit,
        )
        if not excel_ok:
            raise RuntimeError(
                f"Excel ile KPI şablonu doldurulamadı: {excel_mesaj}\n"
                "Pivotlu şablon openpyxl ile güvenle kaydedilemez; Excel kurulu ve dosya kapalı olmalı."
            )
    else:
        print(
            "  Uyarı: Excel COM kullanılamıyor — openpyxl yedek modu. "
            "Pivotlu şablon Excel'de açılamayabilir."
        )
        veri_adet, filo_adet = _openpyxl_yedek_yaz(
            hedef,
            veri_satirlari,
            filo_satirlari,
            veri_sayfa_adlari,
            filo_sayfa_adlari,
            veri_baslik_satiri,
            filo_baslik_satiri,
        )

    print(f"BAŞARILI: KPI şablon raporu → {hedef.resolve()}")
    print(f"  Dönem: {bas} — {bit}")
    print(f"  VERİ satırı: {veri_adet}")
    print(f"  Filo Detay satırı: {filo_adet}")
    if eslesmeyen_veri:
        print(f"  Bilgi: {len(eslesmeyen_veri)} VERİ kolonu şablonda var, Oracle sorgusunda yok (boş kalır — pivotlar etkilenmeyebilir).")
        if getattr(ayarlar, "KPI_ESLEME_UYARISI_DETAY", False):
            ornek = ", ".join(eslesmeyen_veri[:8])
            fazla = len(eslesmeyen_veri) - 8
            ek = f" (+{fazla} kolon daha)" if fazla > 0 else ""
            print(f"    Örnek: {ornek}{ek}")
            print("    → python kpi_sablon_kolon_kesif.py  veya  ayarlar.py KPI_KOLON_ESLEME")
    if excel_mesaj:
        print(f"  Uyarı: {excel_mesaj}")

    return str(hedef.resolve())
