"""Şablondaki VERİ kolon başlıklarını Oracle kolonlarıyla karşılaştırır.

Kullanım (KPI klasöründen):
  python kpi_sablon_kolon_kesif.py
"""
from __future__ import annotations

from openpyxl import load_workbook

import ayarlar
from kpi_sablon_rapor import (
    _kolon_esleme,
    _sayfa_bul,
    sablon_yolu,
)
from kpi_veri import veri_satirlari_getir, veri_semasi_hazir
from oracle_baglanti import baglanti_yonet


def main() -> None:
    sablon = sablon_yolu()
    if not sablon.exists():
        print(f"HATA: Şablon yok → {sablon}")
        return

    veri_adlari = getattr(ayarlar, "KPI_VERI_SAYFA_ADLARI", ["VERİ", "VERI", "Veri"])
    baslik_satiri = int(getattr(ayarlar, "KPI_VERI_BASLIK_SATIRI", 1))

    wb = load_workbook(sablon, read_only=True, data_only=True)
    ws = _sayfa_bul(wb, veri_adlari)
    if ws is None:
        print(f"HATA: VERİ sayfası bulunamadı. Aranan: {veri_adlari}")
        wb.close()
        return

    max_col = ws.max_column or 1
    basliklar = [ws.cell(row=baslik_satiri, column=c).value for c in range(1, max_col + 1)]
    while basliklar and basliklar[-1] is None:
        basliklar.pop()
    wb.close()

    oracle_anahtarlari: list[str] = []
    bas = getattr(ayarlar, "KPI_BASLANGIC_TARIHI", "01.01.2026")
    bit = getattr(ayarlar, "KPI_BITIS_TARIHI", "31.01.2026")
    bind = {"bas": bas, "bit": bit}
    if veri_semasi_hazir()[0]:
        try:
            with baglanti_yonet() as conn:
                satirlar = veri_satirlari_getir(conn.cursor(), bas, bit, bind)
                if satirlar:
                    oracle_anahtarlari = list(satirlar[0].keys())
        except Exception as exc:
            print(f"Oracle okunamadı: {exc}")
    if not oracle_anahtarlari:
        oracle_anahtarlari = [
            "YUK_NO", "YUK_TARIHI", "SEVK_NO", "SEVK_TARIHI", "PROJE_KODU", "PLAKA",
            "ARAC_TIPI", "MUSTERI_KODU", "MUSTERI_ADI", "SUBE_KODU", "SUBE",
            "SATIS_TUTAR", "ALIS_TUTAR", "KAR_ZARAR", "MARJ_YUZDE", "TOPLAM_SATIS",
        ]

    esleme = _kolon_esleme(basliklar, oracle_anahtarlari)

    print(f"Şablon: {sablon}")
    print(f"VERİ kolon sayısı: {len(basliklar)}")
    print(f"Eşleşen: {len(esleme)} / Oracle alan: {len(oracle_anahtarlari)}\n")
    print(f"{'#':>3}  {'Şablon başlığı':<35}  {'Oracle kolon':<20}  Durum")
    print("-" * 75)

    for idx, baslik in enumerate(basliklar, 1):
        if baslik is None or str(baslik).strip() == "":
            continue
        oracle = esleme.get(idx, "—")
        durum = "OK" if idx in esleme else "EKSİK"
        print(f"{idx:3}  {str(baslik)[:35]:<35}  {str(oracle):<20}  {durum}")

    eksik = [str(basliklar[i - 1]) for i in range(1, len(basliklar) + 1) if basliklar[i - 1] and i not in esleme]
    if eksik:
        print("\nayarlar.py örneği:")
        print("KPI_KOLON_ESLEME = {")
        for b in eksik[:10]:
            print(f'    "{b}": "ORACLE_KOLON_ADI",  # TODO')
        print("}")


if __name__ == "__main__":
    main()
