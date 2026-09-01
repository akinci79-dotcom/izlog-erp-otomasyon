"""
Üst yönetim KPI Excel raporu oluşturucu.

Kullanım (KPI klasöründen):
  cd KPI
  python kpi_rapor_olustur.py              # Oracle'dan canlı veri
  python kpi_rapor_olustur.py --ornek      # Örnek veri ile şablon testi
"""
from __future__ import annotations

import os
import sys
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import ayarlar
from kpi_analiz import KpiAnalizSonucu, kpi_analizi_yap, ornek_analiz_sonucu

_KPI_KOKU = Path(__file__).resolve().parent


def _raporlar_klasoru() -> Path:
    klasor = _KPI_KOKU / "raporlar"
    klasor.mkdir(exist_ok=True)
    return klasor


def _kpi_rapor_yolu(dosya_adi: str | None = None) -> Path:
    dosya = dosya_adi or getattr(ayarlar, "KPI_RAPOR_DOSYASI", "kpi_rapor.xlsx")
    if os.path.isabs(dosya):
        return Path(dosya)
    return _raporlar_klasoru() / dosya


# --- Stil sabitleri ---
BASLIK_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
ALT_BASLIK_FONT = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=10)
KPI_FONT = Font(name="Calibri", size=18, bold=True, color="1F4E79")
KPI_ETIKET = Font(name="Calibri", size=9, color="666666")

MAVI_DOLGU = PatternFill("solid", fgColor="1F4E79")
ACIK_MAVI = PatternFill("solid", fgColor="D6E4F0")
YESIL_DOLGU = PatternFill("solid", fgColor="E2EFDA")
SARI_DOLGU = PatternFill("solid", fgColor="FFF2CC")
KIRMIZI_DOLGU = PatternFill("solid", fgColor="FCE4D6")

INce_KENAR = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _format_para(deger) -> str:
    if deger is None:
        return "-"
    if isinstance(deger, Decimal):
        deger = float(deger)
    return f"{deger:,.2f} ₺".replace(",", "X").replace(".", ",").replace("X", ".")


def _format_sayi(deger) -> str:
    if deger is None:
        return "-"
    return f"{int(deger):,}".replace(",", ".")


def _hucre_stil(hucre, font=NORMAL_FONT, fill=None, align="left"):
    hucre.font = font
    hucre.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    hucre.border = INce_KENAR
    if fill:
        hucre.fill = fill


def _baslik_satiri(ws, satir, metin, birlestir=(1, 6)):
    ws.merge_cells(
        start_row=satir,
        start_column=birlestir[0],
        end_row=satir,
        end_column=birlestir[1],
    )
    hucre = ws.cell(row=satir, column=birlestir[0], value=metin)
    _hucre_stil(hucre, font=BASLIK_FONT, fill=MAVI_DOLGU)
    for col in range(birlestir[0], birlestir[1] + 1):
        ws.cell(row=satir, column=col).fill = MAVI_DOLGU


def _tablo_basliklari(ws, satir, basliklar):
    for col, baslik in enumerate(basliklar, 1):
        hucre = ws.cell(row=satir, column=col, value=baslik)
        _hucre_stil(hucre, font=ALT_BASLIK_FONT, fill=ACIK_MAVI)


def _kpi_kutusu(ws, satir, col, etiket, deger, uyari=False):
    etiket_hucre = ws.cell(row=satir, column=col, value=etiket)
    _hucre_stil(etiket_hucre, font=KPI_ETIKET, align="center")
    deger_hucre = ws.cell(row=satir + 1, column=col, value=deger)
    fill = KIRMIZI_DOLGU if uyari else YESIL_DOLGU
    _hucre_stil(deger_hucre, font=KPI_FONT, fill=fill, align="center")
    ws.column_dimensions[get_column_letter(col)].width = 22


def yonetici_ozeti_sayfasi(wb, analiz: KpiAnalizSonucu):
    ws = wb.active
    ws.title = "Yönetici Özeti"
    ozet = analiz.ozet

    _baslik_satiri(ws, 1, "İZLOG LOJİSTİK — ÜST YÖNETİM KPI ÖZETİ")
    ws.cell(row=2, column=1, value=f"Dönem: {ozet.get('donem', '')}  |  Rapor: {ozet.get('rapor_tarihi', '')}")
    ws.merge_cells("A2:F2")

    # KPI kutuları — satır 4
    _kpi_kutusu(ws, 4, 1, "Toplam Yük", _format_sayi(ozet.get("yuk_sayisi")))
    _kpi_kutusu(ws, 4, 2, "Toplam Sevk", _format_sayi(ozet.get("sevk_sayisi")))
    sevk_orani = ozet.get("sevk_orani_yuzde", 0)
    _kpi_kutusu(
        ws, 4, 3, "Sevk Oranı",
        f"%{sevk_orani}",
        uyari=sevk_orani < 90,
    )
    _kpi_kutusu(ws, 4, 4, "Satış Geliri", _format_para(ozet.get("toplam_satis_geliri")))
    _kpi_kutusu(ws, 4, 5, "Yük Başına Ort. Gelir", _format_para(ozet.get("yuk_basina_ortalama_gelir")))

    fatura_orani = ozet.get("fatura_baglama_orani_yuzde", 0)
    _kpi_kutusu(
        ws, 4, 6, "Fatura Bağlama Oranı",
        f"%{fatura_orani}",
        uyari=fatura_orani < 95,
    )

    # Marj satırı
    marj = analiz.marj_analizi
    if marj.get("mevcut"):
        _kpi_kutusu(ws, 7, 1, "Toplam Sevk Alış", _format_para(marj.get("toplam_alis")))
        _kpi_kutusu(ws, 7, 2, "Brüt Marj", _format_para(marj.get("brut_marj")))
        marj_orani = marj.get("brut_marj_orani_yuzde", 0)
        _kpi_kutusu(
            ws, 7, 3, "Brüt Marj Oranı",
            f"%{marj_orani}",
            uyari=marj_orani < 10,
        )

    # Fatura gecikmesi
    fg = analiz.fatura_sagligi
    if fg.get("ort_gecikme_gun") is not None:
        _kpi_kutusu(
            ws, 7, 4, "Ort. Fatura Gecikmesi",
            f"{fg['ort_gecikme_gun']} gün",
            uyari=fg["ort_gecikme_gun"] > 7,
        )

    # Odaklanılması gereken problemler
    satir = 11
    _baslik_satiri(ws, satir, "ODAKLANILMASI GEREKEN PROBLEMLER", birlestir=(1, 6))
    satir += 1
    _tablo_basliklari(ws, satir, ["Öncelik", "Kategori", "Problem", "Detay", "Önerilen Aksiyon"])
    satir += 1

    if not analiz.problemler:
        ws.cell(row=satir, column=1, value="Kritik problem tespit edilmedi.")
        ws.merge_cells(start_row=satir, start_column=1, end_row=satir, end_column=5)
        satir += 1
    else:
        oncelik_renk = {
            "YUKSEK": KIRMIZI_DOLGU,
            "ORTA": SARI_DOLGU,
            "DUSUK": YESIL_DOLGU,
        }
        for problem in analiz.problemler:
            fill = oncelik_renk.get(problem["oncelik"], None)
            for col, anahtar in enumerate(
                ["oncelik", "kategori", "baslik", "detay", "aksiyon"], 1
            ):
                hucre = ws.cell(row=satir, column=col, value=problem.get(anahtar, ""))
                _hucre_stil(hucre, fill=fill)
            satir += 1

    # Uyarılar
    if analiz.uyarilar:
        satir += 1
        _baslik_satiri(ws, satir, "UYARILAR", birlestir=(1, 6))
        satir += 1
        for uyari in analiz.uyarilar:
            ws.cell(row=satir, column=1, value=f"• {uyari}")
            ws.merge_cells(start_row=satir, start_column=1, end_row=satir, end_column=6)
            satir += 1

    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 40


def _veri_sayfasi(wb, baslik, sutunlar, satirlar, deger_formatlayici=None):
    ws = wb.create_sheet(baslik)
    _baslik_satiri(ws, 1, baslik.upper(), birlestir=(1, len(sutunlar)))
    _tablo_basliklari(ws, 3, sutunlar)
    for i, satir in enumerate(satirlar, 4):
        for j, anahtar in enumerate(sutunlar, 1):
            if isinstance(satir, dict):
                deger = satir.get(anahtar)
            else:
                deger = satir[j - 1] if j - 1 < len(satir) else None

            if deger_formatlayici and anahtar in deger_formatlayici:
                deger = deger_formatlayici[anahtar](deger)
            elif isinstance(deger, Decimal):
                deger = float(deger)

            hucre = ws.cell(row=i, column=j, value=deger)
            _hucre_stil(hucre)
    for col in range(1, len(sutunlar) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    return ws


def rapor_olustur(analiz: KpiAnalizSonucu, dosya_adi: str | Path | None = None) -> str:
    dosya = Path(dosya_adi) if dosya_adi else _kpi_rapor_yolu()
    dosya.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    yonetici_ozeti_sayfasi(wb, analiz)

    if analiz.aylik_trend:
        _veri_sayfasi(
            wb,
            "Aylık Trend",
            ["AY", "YUK_SAYISI", "SATIS_GELIRI"],
            analiz.aylik_trend,
            {"SATIS_GELIRI": lambda v: float(v) if v else 0},
        )

    if analiz.proje_performans:
        _veri_sayfasi(
            wb,
            "Proje Performansı",
            ["PROJE_KODU", "YUK_SAYISI", "SATIS_GELIRI", "GELIR_PAYI_YUZDE"],
            analiz.proje_performans,
            {"SATIS_GELIRI": lambda v: float(v) if v else 0},
        )

    if analiz.operasyon_dagilimi:
        _veri_sayfasi(
            wb,
            "Operasyon Dağılımı",
            ["OPERASYON_KODU", "SATIR_SAYISI", "TOPLAM_TUTAR", "GELIR_PAYI_YUZDE"],
            analiz.operasyon_dagilimi,
            {"TOPLAM_TUTAR": lambda v: float(v) if v else 0},
        )

    wb.save(dosya)
    return str(dosya.resolve())


def main():
    ornek_mod = "--ornek" in sys.argv

    if ornek_mod:
        print("Örnek veri ile KPI raporu oluşturuluyor...")
        analiz = ornek_analiz_sonucu()
        dosya = _kpi_rapor_yolu("kpi_rapor_ORNEK.xlsx")
    else:
        print("Oracle'dan KPI verileri çekiliyor...")
        analiz = kpi_analizi_yap()
        dosya = _kpi_rapor_yolu()

    yol = rapor_olustur(analiz, dosya)
    print(f"BAŞARILI: KPI raporu oluşturuldu → {yol}")
    print(f"  Dönem: {analiz.ozet.get('donem')}")
    print(f"  Tespit edilen problem: {len(analiz.problemler)}")


if __name__ == "__main__":
    main()
